import sys
import cv2
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Optional, Tuple
from enum import Enum, auto

from PySide6.QtCore import Qt, QTimer, QThread, Signal
from PySide6.QtGui import QImage, QPixmap, QColor
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QSlider, QTableWidget, QTableWidgetItem,
    QHeaderView, QFileDialog, QDoubleSpinBox, QGroupBox, QMessageBox,
    QSplitter, QSizePolicy, QProgressBar, QFrame
)
from ultralytics import YOLO

# ═══════════════════════════════════════════════════════════════
# 数据层：不可变事件溯源
# ═══════════════════════════════════════════════════════════════

class ActionStatus(Enum):
    NORMAL = auto()
    ABNORMAL = auto()

@dataclass(frozen=True)
class ActionEvent:
    index: int
    start_time: float
    end_time: float
    duration: float
    status: ActionStatus
    start_frame: int
    end_frame: int

    def to_dict(self) -> dict:
        return {
            'action_index': self.index,
            'start_time': self.start_time,
            'end_time': self.end_time,
            'duration': self.duration,
            'status': '异常' if self.status == ActionStatus.ABNORMAL else '正常',
            'start_frame': self.start_frame,
            'end_frame': self.end_frame
        }

@dataclass
class AnalysisResult:
    events: List[ActionEvent] = field(default_factory=list)
    product_count: int = 0
    total_count: int = 0
    abnormal_count: int = 0
    normal_avg_duration: float = 0.0

    @property
    def normal_events(self) -> List[ActionEvent]:
        return [e for e in self.events if e.status == ActionStatus.NORMAL]

    @property
    def abnormal_events(self) -> List[ActionEvent]:
        return [e for e in self.events if e.status == ActionStatus.ABNORMAL]

    @property
    def top10_slowest(self) -> List[ActionEvent]:
        return sorted(self.events, key=lambda x: x.duration, reverse=True)[:10]

    @property
    def top10_fastest(self) -> List[ActionEvent]:
        normals = self.normal_events
        return sorted(normals, key=lambda x: x.duration)[:10]

    def recompute_stats(self):
        self.total_count = len(self.events)
        self.abnormal_count = len(self.abnormal_events)
        normals = self.normal_events
        self.normal_avg_duration = np.mean([e.duration for e in normals]) if normals else 0.0

# ═══════════════════════════════════════════════════════════════
# 核心引擎：纯函数式状态机
# ═══════════════════════════════════════════════════════════════

class StateMachine:
    def __init__(self, confirm_frames: int = 3, cooldown_sec: float = 1.5, 
                 max_cycle_sec: float = 60.0):
        self.confirm_frames = confirm_frames
        self.cooldown_sec = cooldown_sec
        self.max_cycle_sec = max_cycle_sec
        self.reset()

    def reset(self):
        self.state = 'IDLE'
        self.frame_counter = 0
        self.count = 0
        self.last_action_start_time = None
        self.last_action_start_frame = 0
        self.cooldown_start_time = 0.0
        self.events = []
        self._max_processed_frame = -1

    def process(self, detected: bool, time_sec: float, frame_idx: int):
        if frame_idx <= self._max_processed_frame:
            return self.state, None

        event = None

        if self.state == 'IDLE':
            if detected:
                self.frame_counter += 1
                if self.frame_counter >= self.confirm_frames:
                    if self.last_action_start_time is not None:
                        interval = time_sec - self.last_action_start_time
                        status = ActionStatus.ABNORMAL if interval > self.max_cycle_sec else ActionStatus.NORMAL
                        event = ActionEvent(
                            index=self.count,
                            start_time=self.last_action_start_time,
                            end_time=time_sec,
                            duration=interval,
                            status=status,
                            start_frame=self.last_action_start_frame,
                            end_frame=frame_idx
                        )
                        self.events.append(event)

                    self.count += 1
                    self.state = 'LOCKED'
                    self.last_action_start_time = time_sec
                    self.last_action_start_frame = frame_idx
                    self.frame_counter = 0
            else:
                self.frame_counter = 0

        elif self.state == 'LOCKED':
            if not detected:
                self.state = 'COOLDOWN'
                self.cooldown_start_time = time_sec

        elif self.state == 'COOLDOWN':
            if detected:
                self.state = 'LOCKED'
            elif (time_sec - self.cooldown_start_time) >= self.cooldown_sec:
                self.state = 'IDLE'

        self._max_processed_frame = frame_idx
        return self.state, event

    def get_result(self) -> AnalysisResult:
        result = AnalysisResult(events=list(self.events), product_count=self.count)
        result.recompute_stats()
        return result

# ═══════════════════════════════════════════════════════════════
# 视频元数据
# ═══════════════════════════════════════════════════════════════

class VideoMetadata:
    def __init__(self, path: str):
        self.path = path
        cap = cv2.VideoCapture(path)
        if not cap.isOpened():
            raise ValueError(f"无法打开视频: {path}")

        self.fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
        if self.fps <= 0: self.fps = 30.0
        
        self.total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        self.width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        self.height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self.duration_sec = self.total_frames / self.fps
        cap.release()

    def format_time(self, seconds: float) -> str:
        mins = int(seconds // 60)
        secs = int(seconds % 60)
        return f"{mins:02d}:{secs:02d}"

# ═══════════════════════════════════════════════════════════════
# 预处理线程
# ═══════════════════════════════════════════════════════════════

class PreprocessThread(QThread):
    progress = Signal(int, str, object)
    finished = Signal(object)
    error = Signal(str)

    def __init__(self, video_path: str, model_path: str, 
                 confirm_frames: int = 3, cooldown_sec: float = 1.5, 
                 max_cycle_sec: float = 60.0, batch_size: int = 32):
        super().__init__()
        self.video_path = video_path
        self.model_path = model_path
        self.batch_size = batch_size
        self.sm = StateMachine(confirm_frames, cooldown_sec, max_cycle_sec)
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    def run(self):
        cap = cv2.VideoCapture(self.video_path)
        try:
            model = YOLO(self.model_path)
            meta = VideoMetadata(self.video_path)

            batch_frames = []
            batch_indices = []
            idx = 0

            while not self._is_cancelled:
                ret, frame = cap.read()
                if not ret and not batch_frames:
                    break

                if ret:
                    batch_frames.append(frame)
                    batch_indices.append(idx)
                    idx += 1

                if len(batch_frames) == self.batch_size or (not ret and batch_frames):
                    results = model.predict(batch_frames, conf=0.5, verbose=False, batch=self.batch_size)

                    for i, res in enumerate(results):
                        detected = len(res.boxes) > 0
                        time_sec = batch_indices[i] / meta.fps
                        self.sm.process(detected, time_sec, batch_indices[i])

                    progress = int((idx / meta.total_frames) * 100) if meta.total_frames else 100
                    time_text = f"{meta.format_time(idx/meta.fps)} / {meta.format_time(meta.duration_sec)}"
                    self.progress.emit(progress, time_text, self.sm.get_result())

                    batch_frames.clear()
                    batch_indices.clear()

            if not self._is_cancelled:
                self.finished.emit(self.sm.get_result())

        except Exception as e:
            self.error.emit(str(e))
        finally:
            if cap.isOpened():
                cap.release()

# ═══════════════════════════════════════════════════════════════
# 图表生成器
# ═══════════════════════════════════════════════════════════════

class ChartGenerator:
    @staticmethod
    def setup_chinese_font():
        font_paths = [
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/usr/share/fonts/truetype/msyh/msyh.ttc",
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "/System/Library/Fonts/PingFang.ttc",
            "/Library/Fonts/Arial Unicode.ttf"
        ]
        for fp in font_paths:
            if Path(fp).exists():
                fm.fontManager.addfont(fp)
                font_name = fm.FontProperties(fname=fp).get_name()
                plt.rcParams['font.sans-serif'] = [font_name] + plt.rcParams.get('font.sans-serif', [])
                break
        plt.rcParams['axes.unicode_minus'] = False

    @classmethod
    def generate_report_chart(cls, result: AnalysisResult, path: str):
        cls.setup_chinese_font()

        fig, axes = plt.subplots(2, 1, figsize=(10, 6), dpi=150, facecolor='white')
        report_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        fig.suptitle(f'工业视觉效能透视分析报告 (生成于: {report_time})', fontsize=15, color='#333333', weight='bold', y=0.98)

        try:
            ax1, ax2 = axes
            df = pd.DataFrame([e.to_dict() for e in result.events])

            COLOR_NORMAL = '#2E86AB'
            COLOR_ABNORMAL = '#E84855'
            COLOR_AVG_LINE = '#F18F01'

            ax1.set_facecolor('#FAFAFA')
            if not df.empty:
                normal = df[df['status'] == '正常']
                abnormal = df[df['status'] == '异常']

                if not normal.empty:
                    ax1.plot(normal['action_index'], normal['duration'], color=COLOR_NORMAL, marker='o', markersize=3, label='正常节拍', alpha=0.8)
                if not abnormal.empty:
                    ax1.scatter(abnormal['action_index'], abnormal['duration'], color=COLOR_ABNORMAL, s=50, marker='v', label='异常超时', zorder=5, edgecolors='black', linewidths=0.5)

            ax1.set_title('动作耗时趋势监控', color='black', fontsize=11, pad=8)
            ax1.set_xlabel('动作序号', color='#555555')
            ax1.set_ylabel('耗时(秒)', color='#555555')
            ax1.legend(facecolor='white', edgecolor='#CCCCCC', fontsize=9)
            ax1.grid(True, linestyle='--', alpha=0.4, color='#CCCCCC')

            ax2.set_facecolor('#FAFAFA')
            if not df.empty:
                max_t = int(df['duration'].max()) + 2
                bin_width = max(1, int(max_t / 15))
                bins = list(range(0, max_t + bin_width, bin_width))
                ax2.hist(df['duration'], bins=bins, color=COLOR_NORMAL, edgecolor='white', alpha=0.7)

                normal_durations = df[df['status'] == '正常']['duration']
                if not normal_durations.empty:
                    avg = normal_durations.mean()
                    ax2.axvline(avg, color=COLOR_AVG_LINE, linestyle='--', linewidth=2, label=f'正常均值: {avg:.2f}秒')

                ax2.set_title('节拍耗时频率分布', color='black', fontsize=11, pad=8)
                ax2.set_xlabel('耗时区间(秒)', color='#555555')
                ax2.set_ylabel('频次', color='#555555')
                ax2.legend(facecolor='white', edgecolor='#CCCCCC', fontsize=9)
                ax2.grid(True, linestyle='--', alpha=0.4, color='#CCCCCC')

            plt.tight_layout(rect=[0, 0, 1, 0.95])
            plt.savefig(path, facecolor='white', dpi=150, bbox_inches='tight')
        finally:
            plt.close(fig)

# ═══════════════════════════════════════════════════════════════
# 主界面 - 全息未来感重构
# ═══════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):
    F_UI = '"Noto Sans CJK SC", "WenQuanYi Micro Hei", "Microsoft YaHei", sans-serif'
    F_MO = '"Noto Sans Mono CJK SC", "Consolas", monospace'

    def __init__(self):
        super().__init__()
        self.setWindowTitle("工业视觉效能透视终端 V5.0 — 全息空间")

        sg = QApplication.primaryScreen().availableGeometry()
        w = min(2560, max(1200, int(sg.width() * 0.9)))
        h = min(1440, max(800, int(sg.height() * 0.9)))
        self.resize(w, h)
        self.setMinimumSize(1000, 600)

        self._apply_holographic_styles()

        self.model = None
        self.video_meta = None
        self.cap = None
        self.analysis_result = None
        self.preprocess_thread = None
        self.current_result = None

        self.timer = QTimer()
        self.timer.timeout.connect(self._update_frame)
        self.is_playing = False
        self.is_preprocessing = False

        self.replay_sm = StateMachine()

        self._init_ui()
        self._load_model()

    def _apply_holographic_styles(self):
        style = f"""
            QMainWindow {{
                background-color: #050810;
                background-image: qradialgradient(cx:0.5, cy:0.5, radius: 1.5, 
                    fx:0.5, fy:0.5, stop:0 #0a1525, stop:1 #02040a);
            }}
            QWidget#GlassPanel {{
                background-color: rgba(10, 20, 35, 0.65);
                border: 1px solid rgba(0, 240, 255, 0.15);
                border-radius: 8px;
            }}
            QSplitter::handle {{ background-color: transparent; }}
            QSplitter::handle:horizontal {{ width: 2px; margin: 4px 1px; 
                border-left: 1px solid rgba(0, 240, 255, 0.1); border-right: 1px solid rgba(0, 240, 255, 0.1); }}
            QSplitter::handle:vertical {{ height: 2px; margin: 1px 4px; 
                border-top: 1px solid rgba(0, 240, 255, 0.1); border-bottom: 1px solid rgba(0, 240, 255, 0.1); }}
            
            QPushButton {{
                background-color: rgba(0, 240, 255, 0.05); color: #00f0ff;
                border: 1px solid rgba(0, 240, 255, 0.4); padding: 8px 18px;
                font-size: 13px; font-weight: bold; border-radius: 4px; font-family: {self.F_UI};
            }}
            QPushButton:hover {{ background-color: rgba(0, 240, 255, 0.15); border: 1px solid #00f0ff; color: #ffffff; }}
            QPushButton:pressed {{ background-color: rgba(0, 240, 255, 0.3); color: #050810; }}
            QPushButton:disabled {{ background-color: #0d1117; color: #2a3a4a; border-color: #1e2d3d; }}
            
            QSlider::groove:horizontal {{ height: 4px; background: rgba(255, 255, 255, 0.08); border-radius: 2px; }}
            QSlider::handle:horizontal {{ width: 16px; height: 16px; margin: -7px 0;
                background: #050810; border: 2px solid #00f0ff; border-radius: 8px; box-shadow: 0 0 5px #00f0ff; }}
            QSlider::sub-page:horizontal {{ background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #005f6b, stop:1 #00f0ff);
                border-radius: 2px; box-shadow: 0 0 8px #00f0ff; }}
            
            QDoubleSpinBox {{
                background-color: rgba(13, 17, 23, 0.8); color: #00ff9d;
                border: 1px solid rgba(0, 240, 255, 0.2); border-radius: 3px; padding: 4px; font-family: {self.F_MO};
            }}
            
            QTableWidget {{
                background-color: transparent; color: #c9d1d9; gridline-color: rgba(255, 255, 255, 0.05);
                border: none; font-family: {self.F_MO}; alternate-background-color: rgba(0, 240, 255, 0.03);
                selection-background-color: rgba(0, 240, 255, 0.1); selection-color: #00f0ff;
            }}
            QTableWidget::item {{ background-color: transparent; border-bottom: 1px solid rgba(0, 240, 255, 0.05); padding: 8px 4px; }}
            QTableWidget::item:hover {{ background-color: rgba(0, 240, 255, 0.08); }}
            QHeaderView::section {{ background-color: transparent; color: #58a6ff;
                border: none; border-bottom: 1px solid rgba(0, 240, 255, 0.3); 
                font-weight: bold; font-family: {self.F_MO}; padding: 8px 4px;
            }}
            
            QGroupBox {{
                border: 1px solid rgba(0, 240, 255, 0.15); margin-top: 12px; padding-top: 16px;
                color: #58a6ff; font-weight: bold; border-radius: 4px; background-color: rgba(13, 20, 35, 0.4);
            }}
            QGroupBox::title {{ subcontrol-origin: margin; left: 12px; padding: 0 6px; background-color: transparent; }}
            
            QProgressBar {{ border: 1px solid rgba(0, 240, 255, 0.2); border-radius: 3px; text-align: center;
                background: rgba(13, 17, 23, 0.8); color: #00f0ff; height: 20px; }}
            QProgressBar::chunk {{ background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #005f6b, stop:1 #00f0ff);
                border-radius: 2px; box-shadow: 0 0 10px #00f0ff; }}
            QLabel {{ color: #e0e6ed; background: transparent; }}
        """
        self.setStyleSheet(style)

    def _init_ui(self):
        cw = QWidget()
        self.setCentralWidget(cw)
        root = QVBoxLayout(cw)
        sp = 10 if self.width() >= 1920 else 6
        mg = 16 if self.width() >= 1920 else 10
        root.setSpacing(sp)
        root.setContentsMargins(mg, mg, mg, mg)

        ctrl = QHBoxLayout()
        ctrl.setSpacing(10)

        self.btn_load = QPushButton("📁 导入视频")
        self.btn_load.setCursor(Qt.PointingHandCursor)
        self.btn_load.clicked.connect(self._load_video)

        self.btn_play = QPushButton("▶ 播放")
        self.btn_play.setCursor(Qt.PointingHandCursor)
        self.btn_play.clicked.connect(self._toggle_play)
        self.btn_play.setEnabled(False)

        self.btn_export = QPushButton("📊 导出报告")
        self.btn_export.setCursor(Qt.PointingHandCursor)
        self.btn_export.clicked.connect(self._export_report)
        self.btn_export.setEnabled(False)

        self.slider = QSlider(Qt.Horizontal)
        self.slider.setEnabled(False)
        self.slider.sliderReleased.connect(self._on_slider_released)
        self.slider.valueChanged.connect(self._on_slider_moved)

        self.lbl_time = QLabel("00:00 / 00:00")
        time_style = f"font-size:14px; color:#00f0ff; font-family:{self.F_MO}; background:rgba(13, 17, 23, 0.8); padding:4px 12px; border:1px solid rgba(0, 240, 255, 0.2); border-radius:3px;"
        self.lbl_time.setStyleSheet(time_style)
        self.lbl_time.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        ctrl.addWidget(self.btn_load)
        ctrl.addWidget(self.btn_play)
        ctrl.addWidget(self.btn_export)
        ctrl.addWidget(self.slider, stretch=1)
        ctrl.addWidget(self.lbl_time)
        root.addLayout(ctrl)

        self.lbl_status = QLabel(" 系统就绪 — 请导入视频开始分析")
        status_style = f"font-size:12px; color:#58a6ff; background:rgba(13, 20, 35, 0.6); padding:6px 12px; border-left:3px solid #58a6ff; font-family:{self.F_MO}; border-radius:2px;"
        self.lbl_status.setStyleSheet(status_style)
        root.addWidget(self.lbl_status)

        self.main_splitter = QSplitter(Qt.Vertical)

        self.video_splitter = QSplitter(Qt.Horizontal)
        self.lbl_original = QLabel("原始画面")
        self.lbl_annotated = QLabel("AI 识别")
        hud_style = f"background:#000; font-size:16px; color:#2a3a4a; font-family:{self.F_MO}; border-top: 2px solid rgba(0, 240, 255, 0.4); border-bottom: 2px solid rgba(0, 240, 255, 0.4); border-left: 1px solid rgba(0, 240, 255, 0.1); border-right: 1px solid rgba(0, 240, 255, 0.1);"
        for lbl in [self.lbl_original, self.lbl_annotated]:
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            lbl.setMinimumSize(320, 180)
            lbl.setStyleSheet(hud_style)
            
        self.video_splitter.addWidget(self.lbl_original)
        self.video_splitter.addWidget(self.lbl_annotated)
        self.video_splitter.setStretchFactor(0, 1)
        self.video_splitter.setStretchFactor(1, 1)
        self.main_splitter.addWidget(self.video_splitter)

        data_widget = QWidget()
        data_layout = QHBoxLayout(data_widget)
        data_layout.setSpacing(sp)
        data_layout.setContentsMargins(0, 4, 0, 0)

        left = QVBoxLayout()
        left.setSpacing(10)

        # 核心数据区：全息反应堆读出器 (纯中文优化)
        stats_frame = QFrame()
        stats_frame.setObjectName("GlassPanel")
        stats_layout = QVBoxLayout(stats_frame)
        stats_layout.setSpacing(12)
        stats_layout.setContentsMargins(12, 18, 12, 12)

        core_widget = QWidget()
        core_layout = QVBoxLayout(core_widget)
        core_layout.setSpacing(2)
        core_layout.setAlignment(Qt.AlignCenter)
        core_widget.setStyleSheet("""
            QWidget {
                background: qradialgradient(cx:0.5, cy:0.5, radius: 0.7, 
                    fx:0.5, fy:0.5, stop:0 rgba(255, 215, 0, 0.15), stop:1 rgba(10, 20, 35, 0));
                border-radius: 6px;
            }
        """)
        
        self.lbl_product_count = QLabel("0")
        self.lbl_product_count.setStyleSheet(f"font-size:52px; font-weight:900; color:#FFD700; font-family:{self.F_MO}; border:none;")
        self.lbl_product_count.setAlignment(Qt.AlignCenter)
        
        lbl_product_tag = QLabel("产 品 总 数")
        lbl_product_tag.setStyleSheet(f"font-size:14px; color:#8b949e; font-family:{self.F_UI}; letter-spacing: 6px; border:none;")
        lbl_product_tag.setAlignment(Qt.AlignCenter)
        
        core_layout.addWidget(self.lbl_product_count)
        core_layout.addWidget(lbl_product_tag)
        stats_layout.addWidget(core_widget)

        sub_stat_widget = QWidget()
        sub_stat_layout = QHBoxLayout(sub_stat_widget)
        sub_stat_layout.setSpacing(8)
        sub_stat_layout.setContentsMargins(0, 6, 0, 0)

        def create_sub_stat(label_text, color):
            w = QWidget()
            w.setStyleSheet(f"""background: rgba(13, 20, 35, 0.5); border: 1px solid {color}30; border-radius: 4px;""")
            lay = QVBoxLayout(w)
            lay.setSpacing(2)
            lay.setContentsMargins(4, 6, 4, 6)
            lay.setAlignment(Qt.AlignCenter)
            
            val = QLabel("0")
            val.setStyleSheet(f"font-size:20px; font-weight:bold; color:{color}; font-family:{self.F_MO};")
            val.setAlignment(Qt.AlignCenter)
            
            tag = QLabel(label_text)
            tag.setStyleSheet(f"font-size:10px; color:#8b949e; letter-spacing: 2px;")
            tag.setAlignment(Qt.AlignCenter)
            
            lay.addWidget(val)
            lay.addWidget(tag)
            return w, val

        w_cycle, self.lbl_count = create_sub_stat("动作周期", "#00ff9d")
        w_abn, self.lbl_abnormal = create_sub_stat("异常次数", "#ff6b6b")
        w_avg, self.lbl_avg = create_sub_stat("平均周期", "#00f0ff")

        sub_stat_layout.addWidget(w_cycle)
        sub_stat_layout.addWidget(w_abn)
        sub_stat_layout.addWidget(w_avg)
        stats_layout.addWidget(sub_stat_widget)

        self.lbl_progress = QLabel("预处理进度: 0%")
        self.lbl_progress.setStyleSheet(f"font-size:14px; color:#ffa500; font-family:{self.F_MO};")
        self.lbl_progress.setVisible(False)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)

        stats_layout.addWidget(self.lbl_progress)
        stats_layout.addWidget(self.progress_bar)
        left.addWidget(stats_frame)

        param_group = QGroupBox("⚙ 阈值参数")
        param_layout = QVBoxLayout()
        param_layout.setSpacing(8)

        for txt, attr, val, tip in [
            ("防抖冷却(秒):", "spin_cooldown", 1.5, "动作确认后的锁定时间"),
            ("异常熔断(秒):", "spin_max_cycle", 60.0, "超过此时长视为异常")
        ]:
            row = QHBoxLayout()
            lb = QLabel(txt)
            lb.setStyleSheet("font-size:12px; color:#8b949e;")
            lb.setToolTip(tip)
            sp_box = QDoubleSpinBox()
            sp_box.setRange(0.1, 3600.0)
            sp_box.setSingleStep(0.1)
            sp_box.setValue(val)
            sp_box.setDecimals(1)
            sp_box.valueChanged.connect(self._update_params)
            row.addWidget(lb)
            row.addWidget(sp_box, stretch=1)
            param_layout.addLayout(row)
            setattr(self, attr, sp_box)

        param_group.setLayout(param_layout)
        left.addWidget(param_group)
        left.addStretch()
        data_layout.addLayout(left, stretch=0)

        table_splitter = QSplitter(Qt.Vertical)

        def create_table(title, color):
            container = QFrame()
            container.setObjectName("GlassPanel")
            vl = QVBoxLayout(container)
            vl.setSpacing(4)
            vl.setContentsMargins(8, 8, 8, 8)

            title_lbl = QLabel(title)
            title_lbl.setStyleSheet(f"font-size:13px; font-weight:bold; color:{color}; border-bottom:1px solid {color}; padding-bottom:4px; background:transparent;")
            vl.addWidget(title_lbl)

            tw = QTableWidget(0, 4)
            tw.setHorizontalHeaderLabels(["排名", "动作序号", "耗时(秒)", "视频进度"])
            # 【核心修复】强制表头绝对居中
            tw.horizontalHeader().setDefaultAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            tw.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            tw.verticalHeader().setVisible(False)
            tw.verticalHeader().setDefaultSectionSize(32) # 增加行高，更透气
            tw.setEditTriggers(QTableWidget.NoEditTriggers)
            tw.setSelectionBehavior(QTableWidget.SelectRows)
            tw.setShowGrid(False)
            tw.setAlternatingRowColors(True)
            tw.cellClicked.connect(self._jump_to_event)
            vl.addWidget(tw)
            return container, tw

        slow_widget, self.table_slow = create_table("🚨 耗时最长 Top 10", "#ff6b6b")
        fast_widget, self.table_fast = create_table("🏆 效率最高 Top 10", "#00ff9d")

        table_splitter.addWidget(slow_widget)
        table_splitter.addWidget(fast_widget)
        table_splitter.setStretchFactor(0, 1)
        table_splitter.setStretchFactor(1, 1)
        data_layout.addWidget(table_splitter, stretch=1)

        self.main_splitter.addWidget(data_widget)
        self.main_splitter.setSizes([int(self.height()*0.55), int(self.height()*0.45)])
        self.main_splitter.setStretchFactor(0, 1)
        self.main_splitter.setStretchFactor(1, 1)
        root.addWidget(self.main_splitter, stretch=1)

    def _update_params(self):
        if hasattr(self, 'replay_sm'):
            self.replay_sm.cooldown_sec = self.spin_cooldown.value()
            self.replay_sm.max_cycle_sec = self.spin_max_cycle.value()

    def _load_model(self):
        paths = [
            "/home/ai/data_disk1/yolo/runs/train/finish_exp1/weights/best.pt",
            "best.pt", "model/best.pt", "weights/best.pt"
        ]
        for p in paths:
            if Path(p).exists():
                try:
                    self.model = YOLO(p)
                    self._set_status(f"模型已加载: {p}", "success")
                    return True
                except Exception:
                    continue
        self._set_status("⚠ 未找到模型文件，请确保 best.pt 在程序目录中", "error")
        return False

    def _set_status(self, text, level="info"):
        colors = {
            "info": ("#58a6ff", "#58a6ff"), "success": ("#00ff9d", "#00ff9d"),
            "warning": ("#ffa500", "#ffa500"), "error": ("#ff6b6b", "#ff6b6b")
        }
        c, bc = colors.get(level, colors["info"])
        self.lbl_status.setText(f" {text}")
        self.lbl_status.setStyleSheet(f"font-size:12px; color:{c}; background:rgba(13, 20, 35, 0.6); padding:6px 12px; border-left:3px solid {bc}; font-family:{self.F_MO}; border-radius:2px;")

    def _load_video(self):
        if not self.model:
            QMessageBox.warning(self, "警告", "模型未加载，无法分析视频")
            return
        if self.is_preprocessing:
            QMessageBox.information(self, "提示", "正在预处理中，请等待完成")
            return

        path, _ = QFileDialog.getOpenFileName(self, "选择视频源", "", "视频文件 (*.mp4 *.avi *.mov *.mkv *.wmv)")
        if not path: return

        if self.is_playing:
            self.timer.stop()
            self.is_playing = False
            self.btn_play.setText("▶ 播放")

        if self.cap:
            self.cap.release()
            self.cap = None

        try:
            self.video_meta = VideoMetadata(path)
        except ValueError as e:
            QMessageBox.critical(self, "错误", str(e))
            return

        self.cap = cv2.VideoCapture(path)
        self.slider.setRange(0, self.video_meta.total_frames - 1)
        self.slider.setValue(0)
        self.slider.setEnabled(False)
        self.btn_play.setEnabled(False)
        self.btn_export.setEnabled(False)

        self.analysis_result = None
        self.current_result = None
        self.replay_sm.reset()
        self._update_stats_ui(AnalysisResult())
        self._clear_tables()

        self.is_preprocessing = True
        self.lbl_original.setText("GPU 加速中...")
        self.lbl_annotated.setText("AI 推理中...")
        self._set_status("正在调用 GPU 极速预处理，请稍候...", "warning")

        self.lbl_progress.setVisible(True)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)

        self.preprocess_thread = PreprocessThread(
            path, self.model.ckpt_path if self.model else "best.pt",
            confirm_frames=3, cooldown_sec=self.spin_cooldown.value(), max_cycle_sec=self.spin_max_cycle.value()
        )
        self.preprocess_thread.progress.connect(self._on_preprocess_progress)
        self.preprocess_thread.finished.connect(self._on_preprocess_finished)
        self.preprocess_thread.error.connect(self._on_preprocess_error)
        self.preprocess_thread.start()

    def _on_preprocess_progress(self, progress, time_text, result):
        self.progress_bar.setValue(progress)
        self.lbl_progress.setText(f"预处理进度: {progress}%")
        self.lbl_time.setText(f"[ {progress}% ] {time_text}")
        self.current_result = result
        self._update_stats_ui(result)

    def _on_preprocess_finished(self, result):
        self.is_preprocessing = False
        self.analysis_result = result
        self.current_result = result
        self.slider.setEnabled(True)
        self.slider.setValue(0)
        self.btn_play.setEnabled(True)
        self.btn_export.setEnabled(True)
        self.lbl_progress.setVisible(False)
        self.progress_bar.setVisible(False)
        self.lbl_original.setText("审查回放模式")
        self.lbl_annotated.setText("安全防重锁已激活")
        self._set_status(f"预处理完成！产品总数: {result.product_count}，异常 {result.abnormal_count} 次", "success")
        self.lbl_time.setText(f"00:00 / {self.video_meta.format_time(self.video_meta.duration_sec)}")
        self._update_stats_ui(result)
        self._update_tables(result)

    def _on_preprocess_error(self, error_msg):
        self.is_preprocessing = False
        self.lbl_progress.setVisible(False)
        self.progress_bar.setVisible(False)
        self._set_status(f"预处理失败: {error_msg}", "error")
        QMessageBox.critical(self, "预处理失败", f"错误详情:\n{error_msg}")

    def _toggle_play(self):
        if not self.cap or self.is_preprocessing: return
        self.is_playing = not self.is_playing
        self.btn_play.setText("⏸ 暂停" if self.is_playing else "▶ 播放")
        if self.is_playing:
            self.timer.start(int(1000 / self.video_meta.fps))
        else:
            self.timer.stop()

    def _update_frame(self):
        if not self.cap or not self.is_playing: return
        ret, frame = self.cap.read()
        if not ret:
            self.timer.stop()
            self.is_playing = False
            self.btn_play.setText("▶ 播放 (已结束)")
            return

        cur_frame = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES))
        cur_time = cur_frame / self.video_meta.fps
        self.slider.setValue(cur_frame)
        self.lbl_time.setText(f"{self.video_meta.format_time(cur_time)} / {self.video_meta.format_time(self.video_meta.duration_sec)}")

        annotated = frame.copy()
        detected = False
        if self.model:
            results = self.model.predict(frame, conf=0.5, verbose=False)
            if results and len(results[0].boxes) > 0:
                detected = True
                annotated = results[0].plot()

        state, _ = self.replay_sm.process(detected, cur_time, cur_frame)
        self._draw_holographic_overlay(annotated, cur_frame, state, detected)
        self._show_frame(frame, annotated)

    def _draw_holographic_overlay(self, img, frame_idx, state, detected):
        count = self.replay_sm.count
        overlay = img.copy()
        cv2.rectangle(overlay, (15, 15), (350, 145), (5, 10, 20), -1)
        cv2.addWeighted(overlay, 0.6, img, 0.4, 0, img)
        
        cv2.putText(img, f"PRODUCT: {count}", (30, 50), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 240, 255), 2, cv2.LINE_AA)

        status_text = f"STATE: {state}"
        color = (0, 255, 157) if state == 'IDLE' else (0, 165, 255) if state == 'LOCKED' else (255, 165, 0)
        cv2.putText(img, status_text, (30, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2, cv2.LINE_AA)

        det_text = "DETECTED" if detected else "NO TARGET"
        det_color = (0, 255, 0) if detected else (128, 128, 128)
        cv2.putText(img, det_text, (30, 130), cv2.FONT_HERSHEY_SIMPLEX, 0.7, det_color, 2, cv2.LINE_AA)

    def _on_slider_moved(self, value):
        if not self.video_meta: return
        t = value / self.video_meta.fps
        self.lbl_time.setText(f"{self.video_meta.format_time(t)} / {self.video_meta.format_time(self.video_meta.duration_sec)}")

    def _on_slider_released(self):
        if not self.cap or not self.video_meta: return
        pos = max(0, min(self.slider.value(), self.video_meta.total_frames - 1))
        self.cap.set(cv2.CAP_PROP_POS_FRAMES, pos)
        self.replay_sm.reset()
        if not self.is_playing:
            ret, frame = self.cap.read()
            if ret:
                annotated = frame.copy()
                if self.model:
                    results = self.model.predict(frame, conf=0.5, verbose=False)
                    if results and len(results[0].boxes) > 0:
                        annotated = results[0].plot()
                self._draw_holographic_overlay(annotated, pos, self.replay_sm.state, False)
                self._show_frame(frame, annotated)

    def _jump_to_event(self, row, col):
        if not self.cap or not self.video_meta: return
        table = self.sender()
        item = table.item(row, 3)
        if not item: return
        try:
            time_str = item.text().split(" - ")[0]
            mins, secs = map(int, time_str.split(":"))
            target_frame = int((mins * 60 + secs) * self.video_meta.fps)
            target_frame = max(0, min(target_frame, self.video_meta.total_frames - 1))
            self.slider.setValue(target_frame)
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, target_frame)
            self.replay_sm.reset()
            ret, frame = self.cap.read()
            if ret:
                annotated = frame.copy()
                if self.model:
                    results = self.model.predict(frame, conf=0.5, verbose=False)
                    if results and len(results[0].boxes) > 0:
                        annotated = results[0].plot()
                self._draw_holographic_overlay(annotated, target_frame, self.replay_sm.state, False)
                self._show_frame(frame, annotated)
        except (ValueError, IndexError): pass

    def _show_frame(self, original, annotated):
        for img, lbl in [(original, self.lbl_original), (annotated, self.lbl_annotated)]:
            rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb.shape
            qt_img = QImage(rgb.data, w, h, ch * w, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qt_img)
            lbl.setPixmap(pixmap.scaled(lbl.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))

    def _update_stats_ui(self, result):
        self.lbl_product_count.setText(f"{result.product_count}")
        self.lbl_count.setText(f"{result.total_count}")
        self.lbl_abnormal.setText(f"{result.abnormal_count}")
        self.lbl_avg.setText(f"{result.normal_avg_duration:.2f}")

    def _update_tables(self, result):
        # --- 耗时最长 Top 10 (问题清单) ---
        self.table_slow.setRowCount(0)
        for i, evt in enumerate(result.top10_slowest):
            self.table_slow.insertRow(i)
            # 修复：剥离文本拼接，动作序号只保留纯数字
            items = [
                str(i + 1),
                str(evt.index), 
                f"{evt.duration:.2f}",
                f"{self.video_meta.format_time(evt.start_time)} - {self.video_meta.format_time(evt.end_time)}"
            ]
            for j, txt in enumerate(items):
                it = QTableWidgetItem(txt)
                # 【核心修复】强制每个单元格数据绝对居中
                it.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                
                # 修复：仅通过颜色和背景区分状态，保持数据纯净
                if evt.status == ActionStatus.ABNORMAL:
                    it.setForeground(QColor("#ff6b6b"))
                    it.setBackground(QColor(255, 107, 107, 25))  # 极淡的红色背景微光
                else:
                    it.setForeground(QColor("#c9d1d9"))  # 正常但偏慢的数据为冷灰白
                
                self.table_slow.setItem(i, j, it)

        # --- 效率最高 Top 10 (参考基准) ---
        self.table_fast.setRowCount(0)
        for i, evt in enumerate(result.top10_fastest):
            self.table_fast.insertRow(i)
            items = [
                str(i + 1), 
                str(evt.index), 
                f"{evt.duration:.2f}",
                f"{self.video_meta.format_time(evt.start_time)} - {self.video_meta.format_time(evt.end_time)}"
            ]
            for j, txt in enumerate(items):
                it = QTableWidgetItem(txt)
                # 【核心修复】强制每个单元格数据绝对居中
                it.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                # 修复：标杆区统一使用柔和的青绿色，暗示平稳高效
                it.setForeground(QColor("#00ff9d")) 
                self.table_fast.setItem(i, j, it)

    def _clear_tables(self):
        self.table_slow.setRowCount(0)
        self.table_fast.setRowCount(0)

    def _export_report(self):
        if not self.analysis_result or not self.analysis_result.events:
            QMessageBox.information(self, "提示", "无数据可导出！")
            return

        default_name = f"效能分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        fp, _ = QFileDialog.getSaveFileName(self, "导出报告", default_name, "Excel 文件 (*.xlsx)")
        if not fp: return

        result = self.analysis_result
        temp_img = str(Path(fp).with_suffix('.png'))

        try:
            ChartGenerator.generate_report_chart(result, temp_img)

            def fmt_sec(sec: float) -> str:
                h, rem = divmod(int(sec), 3600)
                m, s = divmod(rem, 60)
                return f"{h:02d}:{m:02d}:{s:02d}" if h > 0 else f"{m:02d}:{s:02d}"

            detail_rows = []
            abnormal_indices = {i + 2 for i, e in enumerate(result.events) if e.status == ActionStatus.ABNORMAL}
            
            for e in result.events:
                detail_rows.append({
                    '动作序号': e.index, '开始时间': fmt_sec(e.start_time), '结束时间': fmt_sec(e.end_time),
                    '耗时(秒)': e.duration, '状态': '异常' if e.status == ActionStatus.ABNORMAL else '正常',
                    '起始帧': e.start_frame, '结束帧': e.end_frame
                })
            df_detail = pd.DataFrame(detail_rows)

            summary_data = {
                '指标': ['产品总数', '动作周期数', '异常行为次数', '正常节拍数', '正常平均周期(秒)'],
                '数值': [result.product_count, result.total_count, result.abnormal_count, result.total_count - result.abnormal_count, round(result.normal_avg_duration, 2)]
            }
            df_summary = pd.DataFrame(summary_data)

            thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
            center_align = Alignment(horizontal='center', vertical='center')
            header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell_font = Font(name='微软雅黑', size=10)
            err_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            err_font = Font(name='微软雅黑', size=10, color='9C0006', bold=True)

            def get_exact_width(value) -> float:
                if not value: return 6.0
                length = 0.0
                for char in str(value):
                    length += 2.0 if '\u4e00' <= char <= '\u9fff' else 1.1
                return length + 4.0

            with pd.ExcelWriter(fp, engine='openpyxl') as writer:
                df_summary.to_excel(writer, sheet_name='汇总数据', index=False, startrow=1)
                df_detail.to_excel(writer, sheet_name='动作明细', index=False)

                wb = writer.book
                ws_sum = wb['汇总数据']
                ws_sum.merge_cells('A1:B1')
                title_cell = ws_sum['A1']
                title_cell.value = f"效能分析汇总 (生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')})"
                title_cell.font = Font(name='微软雅黑', bold=True, size=13, color='1F4E79')
                title_cell.alignment = Alignment(horizontal='left', vertical='center')
                ws_sum.row_dimensions[1].height = 30

                for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row, max_col=2):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = center_align
                        if cell.row == 2:
                            cell.font = header_font; cell.fill = header_fill
                        else:
                            cell.font = cell_font
                    col_letter = row[0].column_letter
                    ws_sum.column_dimensions[col_letter].width = max(get_exact_width(row[0].value), get_exact_width(row[1].value))

                ws_det = wb['动作明细']
                ws_det.freeze_panes = 'A2'
                for cell in ws_det[1]:
                    ws_det.column_dimensions[cell.column_letter].width = get_exact_width(cell.value)

                for row in ws_det.iter_rows(min_row=2, max_row=ws_det.max_row, max_col=ws_det.max_column):
                    is_err_row = row[0].row in abnormal_indices
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = center_align
                        if is_err_row:
                            cell.fill = err_fill; cell.font = err_font
                        else:
                            cell.font = cell_font

                ws_chart = wb.create_sheet("可视化图表")
                img = XLImage(temp_img)
                img.width = 960; img.height = 580
                ws_chart.add_image(img, 'A1')

            QMessageBox.information(self, "导出成功", f"专业级分析报告已保存至:\n{fp}")

        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"错误详情:\n{str(e)}")
        finally:
            if Path(temp_img).exists():
                Path(temp_img).unlink()

    def closeEvent(self, event):
        if self.preprocess_thread and self.preprocess_thread.isRunning():
            self.preprocess_thread.cancel()
            self.preprocess_thread.wait(3000)
        if self.cap:
            self.cap.release()
        event.accept()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())