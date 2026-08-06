import cv2
import numpy as np
import os
import logging
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import matplotlib.patches as mpatches
from ultralytics import YOLO
from PySide6.QtCore import Qt, QTimer, QThread, Signal
from PySide6.QtGui import QImage, QPixmap, QColor, QShortcut, QKeySequence
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                               QHBoxLayout, QPushButton, QSlider, QLabel, 
                               QFileDialog, QMessageBox, QTableWidget, 
                               QTableWidgetItem, QHeaderView, QProgressBar, QFrame, QSizePolicy)
import sys

os.environ['YOLO_VERBOSE'] = 'False'
logging.getLogger("AuditSystem").setLevel(logging.WARNING)

# ★ 极致强化版 Matplotlib 中文字体配置 ★
def setup_matplotlib_font():
    font_found = False
    # 1. 尝试在系统已注册字体中寻找中文字体
    for font in fm.fontManager.ttflist:
        if any(keyword in font.name.lower() for keyword in ['cjk', 'noto sans', 'simhei', 'msyh', 'wqy', 'pingfang', 'heiti']):
            plt.rcParams['font.sans-serif'] = [font.name] + plt.rcParams.get('font.sans-serif', [])
            font_found = True
            break
    
    # 2. 如果系统缓存里没找到，尝试硬加载常见 Linux 路径
    if not font_found:
        font_paths = [
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/wenquanyi/wqy-zenhei.ttc",
            "/usr/share/fonts/wqy-microhei/wqy-microhei.ttc",
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "/System/Library/Fonts/PingFang.ttc"
        ]
        for fp in font_paths:
            if Path(fp).exists():
                fm.fontManager.addfont(fp)
                font_prop = fm.FontProperties(fname=fp)
                plt.rcParams['font.sans-serif'] = [font_prop.get_name()] + plt.rcParams.get('font.sans-serif', [])
                font_found = True
                break
                
    plt.rcParams['axes.unicode_minus'] = False

setup_matplotlib_font()

# ═══════════════════════════════════════════════════════════════
# 核心引擎：目标锁死追踪状态机 (含瞬时动作过滤)
# ═══════════════════════════════════════════════════════════════
class CausalStateMachine:
    def __init__(self, hard_thresh=17, area_thresh=500, confirm_frames=6, cooldown_frames=10):
        self.reset_state()
        self.confirm_frames = confirm_frames
        self.cooldown_frames = cooldown_frames
        self.hard_thresh = hard_thresh  
        self.area_thresh = area_thresh
        self.frame_idx = 0
        self.clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        
    def reset_state(self):
        self.state = 'IDLE'
        self.intrusion_counter = 0
        self.exit_counter = 0
        self.locked_counter = 0
        self.count = 0
        self.locked_roi = None        
        self.base_roi_gray = None 
        self.trigger_hand_center = None
        self.intrusion_start_frame = 0  
        self.intrusion_duration = 0     

    def _calc_intrusion_ratio(self, hands):
        if not self.locked_roi or not hands: return 0.0
        bx1, by1, bx2, by2 = self.locked_roi
        max_ratio = 0.0
        for hx1, hy1, hx2, hy2 in hands:
            ix1, iy1 = max(hx1, bx1), max(hy1, by1)
            ix2, iy2 = min(hx2, bx2), min(hy2, by2)
            if ix1 < ix2 and iy1 < iy2:
                inter_area = (ix2 - ix1) * (iy2 - iy1)
                hand_area = (hx2 - hx1) * (hy2 - hy1)
                if hand_area > 0:
                    ratio = inter_area / hand_area
                    if ratio > max_ratio: max_ratio = ratio
        return max_ratio

    def process(self, frame, hands, basket_box):
        self.frame_idx += 1
        intrusion_ratio = self._calc_intrusion_ratio(hands)
        
        if self.state == 'IDLE' and intrusion_ratio == 0.0 and basket_box is not None:
            self.locked_roi = basket_box

        if self.locked_roi is None:
            return self.state, 0, None, 0.0

        rx1, ry1, rx2, ry2 = self.locked_roi
        roi_gray = cv2.cvtColor(frame[ry1:ry2, rx1:rx2], cv2.COLOR_BGR2GRAY)
        
        p_diff_area = 0
        crop_patch = None

        if self.state == 'IDLE':
            if intrusion_ratio == 0.0:
                self.base_roi_gray = roi_gray.copy()
            if intrusion_ratio > 0.3:
                self.intrusion_counter += 1
                if self.intrusion_counter >= self.confirm_frames:
                    self.state = 'INTRUSION'
                    self.intrusion_counter = 0
                    self.intrusion_start_frame = self.frame_idx 
                    bx1, by1, bx2, by2 = self.locked_roi
                    best_hand = None
                    max_r = 0.0
                    for h in hands:
                        hx1, hy1, hx2, hy2 = h
                        ix1, iy1 = max(hx1, bx1), max(hy1, by1)
                        ix2, iy2 = min(hx2, bx2), min(hy2, by2)
                        if ix1 < ix2 and iy1 < iy2:
                            r = ((ix2-ix1)*(iy2-iy1)) / ((hx2-hx1)*(hy2-hy1))
                            if r > max_r: max_r = r; best_hand = h
                    if best_hand:
                        self.trigger_hand_center = ((best_hand[0] + best_hand[2]) / 2, (best_hand[1] + best_hand[3]) / 2)
            else:
                self.intrusion_counter = 0

        elif self.state == 'INTRUSION':
            self.intrusion_duration = self.frame_idx - self.intrusion_start_frame 
            target_hand = None
            min_dist = float('inf')
            for h in hands:
                hx1, hy1, hx2, hy2 = h
                cx = (hx1 + hx2) / 2
                cy = (hy1 + hy2) / 2
                dist = (cx - self.trigger_hand_center[0])**2 + (cy - self.trigger_hand_center[1])**2
                if dist < min_dist:
                    min_dist = dist
                    target_hand = h
            if target_hand is None or min_dist > 40000:
                self.exit_counter = 0
                return self.state, 0, None, 0.0
            hx1, hy1, hx2, hy2 = target_hand
            self.trigger_hand_center = ((hx1+hx2)/2, (hy1+hy2)/2)
            bx1, by1, bx2, by2 = self.locked_roi
            ix1, iy1 = max(hx1, bx1), max(hy1, by1)
            ix2, iy2 = min(hx2, bx2), min(hy2, by2)
            target_ratio = 0.0
            if ix1 < ix2 and iy1 < iy2:
                inter_area = (ix2 - ix1) * (iy2 - iy1)
                hand_area = (hx2 - hx1) * (hy2 - hy1)
                if hand_area > 0:
                    target_ratio = inter_area / hand_area
            if target_ratio < 0.02:
                self.exit_counter += 1
                if self.exit_counter >= self.confirm_frames:
                    self.state = 'VERIFICATION'
                    self.exit_counter = 0
            else:
                self.exit_counter = 0

        elif self.state == 'VERIFICATION':
            if self.base_roi_gray is None or self.base_roi_gray.shape != roi_gray.shape:
                self.state = 'IDLE'
                return self.state, 0, None, intrusion_ratio
            
            if self.intrusion_duration < 8: 
                self.state = 'LOCKED'
                return self.state, 0, None, intrusion_ratio

            t0_clahe = self.clahe.apply(self.base_roi_gray)
            t2_clahe = self.clahe.apply(roi_gray)
            t0_blur = cv2.blur(t0_clahe, (31, 31))
            t2_blur = cv2.blur(t2_clahe, (31, 31))
            diff = cv2.absdiff(t0_blur, t2_blur)
            _, p_thresh = cv2.threshold(diff, self.hard_thresh, 255, cv2.THRESH_BINARY)
            p_thresh = cv2.morphologyEx(p_thresh, cv2.MORPH_OPEN, np.ones((5, 5), np.uint8))
            p_thresh = cv2.morphologyEx(p_thresh, cv2.MORPH_CLOSE, np.ones((15, 15), np.uint8))
            contours, _ = cv2.findContours(p_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if contours:
                c = max(contours, key=cv2.contourArea)
                p_diff_area = cv2.contourArea(c)
                if p_diff_area > self.area_thresh:
                    self.count += 1
                    x, y, w, h = cv2.boundingRect(c)
                    layer_mask = np.zeros(roi_gray.shape, dtype=np.uint8)
                    cv2.drawContours(layer_mask, [c], -1, 255, thickness=cv2.FILLED)
                    pure_object = cv2.bitwise_and(frame[ry1:ry2, rx1:rx2], frame[ry1:ry2, rx1:rx2], mask=layer_mask)
                    crop_patch = pure_object[y:y+h, x:x+w].copy()
            self.state = 'LOCKED'

        elif self.state == 'LOCKED':
            if self.base_roi_gray is not None and self.base_roi_gray.shape == roi_gray.shape:
                self.base_roi_gray = roi_gray.copy()
            self.locked_counter += 1
            if self.locked_counter >= self.cooldown_frames:
                self.state = 'IDLE'
                self.locked_counter = 0
        return self.state, int(p_diff_area), crop_patch, intrusion_ratio


# ═══════════════════════════════════════════════════════════════
# 极速后台处理线程
# ═══════════════════════════════════════════════════════════════
class FastProcessThread(QThread):
    progress = Signal(int)
    finished = Signal(list, float, int, str, int, list) 
    error = Signal(str)

    def __init__(self, video_path, model_path):
        super().__init__()
        self.video_path = video_path
        self.model_path = model_path

    def run(self):
        try:
            cap_tmp = cv2.VideoCapture(self.video_path)
            fps = cap_tmp.get(cv2.CAP_PROP_FPS) or 30.0
            total_frames = int(cap_tmp.get(cv2.CAP_PROP_FRAME_COUNT))
            cap_tmp.release()

            yolo = YOLO(self.model_path)
            sm = CausalStateMachine()
            raw_records = []
            last_count = 0
            last_count_frame = 0
            frame_idx = 0
            count_history = [(0, 0)]

            results = yolo.predict(source=self.video_path, stream=True, conf=0.25, iou=0.45, verbose=False)
            for result in results:
                frame_idx += 1
                frame = result.orig_img
                boxes = result.boxes.data.cpu().numpy()
                
                hands = []
                basket_box = None
                for b in boxes:
                    x1, y1, x2, y2 = map(int, b[:4])
                    cls = int(b[5])
                    if cls == 0: hands.append((x1, y1, x2, y2))
                    elif cls == 1: basket_box = (x1, y1, x2, y2)
                
                state, p_diff, crop_patch, ratio = sm.process(frame, hands, basket_box)
                
                if sm.count > last_count:
                    current_frame = frame_idx
                    count_history.append((current_frame, sm.count))
                    if last_count >= 1:
                        duration_frames = current_frame - last_count_frame
                        duration_time = duration_frames / fps
                        raw_records.append({
                            'start_frame': last_count_frame,
                            'end_frame': current_frame,
                            'time': duration_time
                        })
                    last_count_frame = current_frame
                    last_count = sm.count

                if frame_idx % 100 == 0:
                    prog = int((frame_idx / total_frames) * 100) if total_frames > 0 else 0
                    self.progress.emit(prog)

            cycle_records = []
            if len(raw_records) > 2:
                times = [r['time'] for r in raw_records]
                median_time = np.median(times)
                min_valid_time = max(3.0, median_time * 0.4) 
                
                valid_id = 1
                buffer_record = None 
                
                for r in raw_records:
                    if r['time'] < min_valid_time:
                        if buffer_record is not None:
                            buffer_record['end_frame'] = r['end_frame']
                            buffer_record['time'] = (buffer_record['end_frame'] - buffer_record['start_frame']) / fps
                        else:
                            buffer_record = r
                    else:
                        if buffer_record is not None:
                            if buffer_record['time'] >= min_valid_time:
                                buffer_record['cycle_id'] = valid_id
                                cycle_records.append(buffer_record)
                                valid_id += 1
                                buffer_record = r
                            else:
                                r['start_frame'] = buffer_record['start_frame']
                                r['time'] = (r['end_frame'] - r['start_frame']) / fps
                                buffer_record = r
                        else:
                            buffer_record = r
                
                if buffer_record is not None:
                    buffer_record['cycle_id'] = valid_id
                    cycle_records.append(buffer_record)
            else:
                for i, r in enumerate(raw_records):
                    r['cycle_id'] = i + 1
                    cycle_records.append(r)

            self.finished.emit(cycle_records, fps, total_frames, self.video_path, sm.count, count_history)
        except Exception as e:
            self.error.emit(str(e))


# ═══════════════════════════════════════════════════════════════
# 工业工程(IE)统计算法工具
# ═══════════════════════════════════════════════════════════════
def calculate_standard_time(times):
    if len(times) == 0: return 0.0
    median = np.median(times)
    mad = np.median(np.abs(times - median))
    std_est = max(0.5, 1.4826 * mad) 
    valid_times = [t for t in times if abs(t - median) <= 3 * std_est]
    return np.mean(valid_times) if valid_times else median

def calculate_compliance_rate(times, std_time):
    if len(times) == 0 or std_time == 0: return 0.0
    lower_bound = std_time * 0.8
    upper_bound = std_time * 1.2
    compliant_count = len([t for t in times if lower_bound <= t <= upper_bound])
    return (compliant_count / len(times)) * 100

def get_excel_col_width(text):
    if not text: return 0
    length = 0
    for char in str(text):
        if '\u4e00' <= char <= '\u9fff' or char in ['：', '（', '）', '，']:
            length += 2.1
        else:
            length += 1.1
    return length


# ═══════════════════════════════════════════════════════════════
# Demo 全息战术复盘界面
# ═══════════════════════════════════════════════════════════════
class DemoApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("工业视觉效能透视终端 - 全息战术复盘版")
        self.resize(1600, 900)
        
        self.yolo = None
        self.model_path = None
        self.cap = None
        self.fps = 30.0
        self.total_frames = 0
        self.video_path = None
        self.cycle_records = []  
        self.total_count = 0
        self.count_history = [(0, 0)]
        self.process_thread = None

        self._init_ui()
        self._init_shortcuts()
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_frame)
        
    def _init_ui(self):
        self.setStyleSheet("""
            QMainWindow { background-color: #050810; }
            QWidget { color: #e0e6ed; font-family: 'Consolas', 'Microsoft YaHei'; }
            QPushButton {
                background-color: rgba(0, 240, 255, 0.1); color: #00f0ff; border: 1px solid #00f0ff;
                padding: 8px 18px; font-size: 13px; font-weight: bold; border-radius: 4px;
            }
            QPushButton:hover { background-color: rgba(0, 240, 255, 0.3); color: #ffffff; }
            QPushButton:disabled { color: #2a3a4a; border-color: #1e2d3d; background-color: rgba(13, 17, 23, 0.5); }
            QSlider::groove:horizontal { height: 4px; background: rgba(255, 255, 255, 0.1); border-radius: 2px; }
            QSlider::handle:horizontal { width: 14px; height: 14px; margin: -6px 0; background: #00f0ff; border: 2px solid #050810; border-radius: 8px; }
            QSlider::sub-page:horizontal { background: #00f0ff; border-radius: 2px; }
            QTableWidget {
                background-color: rgba(10, 20, 35, 0.8); color: #c9d1d9; gridline-color: rgba(0, 240, 255, 0.1);
                border: 1px solid rgba(0, 240, 255, 0.2); border-radius: 4px;
            }
            QHeaderView::section { background-color: rgba(0, 240, 255, 0.1); color: #00f0ff; border: none; padding: 8px 4px; font-weight: bold; }
            QProgressBar { border: 1px solid rgba(0, 240, 255, 0.3); border-radius: 3px; text-align: center; background: rgba(13, 17, 23, 0.8); color: #00f0ff; }
            QProgressBar::chunk { background-color: #00f0ff; border-radius: 2px; }
            QLabel { background: transparent; }
        """)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # 顶部控制栏
        top_layout = QHBoxLayout()
        self.btn_model = QPushButton("🧠 导入模型")
        self.btn_model.clicked.connect(self.load_model)
        
        self.btn_load = QPushButton("📁 导入视频并分析")
        self.btn_load.clicked.connect(self.load_video)
        self.btn_load.setEnabled(False)
        
        self.btn_export = QPushButton("📊 导出报告")
        self.btn_export.clicked.connect(self.export_report)
        self.btn_export.setEnabled(False)
        
        self.btn_play = QPushButton("▶ 播放")
        self.btn_play.clicked.connect(self.toggle_play)
        self.btn_play.setEnabled(False)
        
        self.slider = QSlider(Qt.Horizontal)
        self.slider.setRange(0, 0)
        self.slider.sliderMoved.connect(self.seek)
        
        self.lbl_time = QLabel("00:00 / 00:00")
        self.lbl_time.setStyleSheet("color: #00f0ff; font-size: 14px; font-weight: bold;")
        
        top_layout.addWidget(self.btn_model)
        top_layout.addWidget(self.btn_load)
        top_layout.addWidget(self.btn_play)
        top_layout.addWidget(self.btn_export)
        top_layout.addWidget(self.slider, stretch=1)
        top_layout.addWidget(self.lbl_time)
        main_layout.addLayout(top_layout)
        
        # 中间视频显示区
        video_frame = QFrame()
        video_frame.setStyleSheet("border: 2px solid rgba(0, 240, 255, 0.3); border-radius: 8px; background-color: black;")
        video_layout = QVBoxLayout(video_frame)
        video_layout.setContentsMargins(0, 0, 0, 0)
        
        self.video_label = QLabel("⚡ 系统就绪 ⚡\n\n请点击 [🧠 导入模型] 加载权重")
        self.video_label.setAlignment(Qt.AlignCenter)
        self.video_label.setStyleSheet("color: #00f0ff; font-size: 22px; font-weight: bold; background-color: transparent;")
        self.video_label.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        self.video_label.setScaledContents(False)
        video_layout.addWidget(self.video_label)
        main_layout.addWidget(video_frame, stretch=3)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("⚡ 正在极速全息分析中... %p%")
        main_layout.addWidget(self.progress_bar)
        
        # 底部看板区
        board_widget = QWidget()
        board_layout = QVBoxLayout(board_widget)
        self.board_title = QLabel("📊 生产节拍效能看板 (待分析)")
        self.board_title.setStyleSheet("font-size: 16px; font-weight: bold; color: #ff9d00; border-bottom: 2px solid rgba(255, 157, 0, 0.3); padding-bottom: 5px;")
        board_layout.addWidget(self.board_title)
        
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["排名", "周期序号", "耗时(秒)", "时间段(起-止)"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.cellClicked.connect(self.on_table_clicked)
        board_layout.addWidget(self.table)
        main_layout.addWidget(board_widget, stretch=1)

    def _init_shortcuts(self):
        QShortcut(QKeySequence(Qt.Key_Space), self, activated=self.toggle_play)
        QShortcut(QKeySequence("Ctrl+E"), self, activated=self.export_report)

    def load_model(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择 YOLO 模型权重", "/home/ai/data_disk1/yolo", "PyTorch Model (*.pt)")
        if not file_path: return
        try:
            self.model_path = file_path
            self.yolo = YOLO(file_path)
            self.btn_load.setEnabled(True)
            self.video_label.setText("✅ 模型加载成功\n\n请点击 [📁 导入视频并分析]")
            QMessageBox.information(self, "成功", f"模型已加载:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"模型加载失败:\n{str(e)}")

    def load_video(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择视频", "/home/ai/data_disk1/yolo/PJ1", "Video Files (*.mp4 *.avi *.mov)")
        if not file_path: return
        
        self.video_path = file_path
        self.btn_load.setEnabled(False)
        self.btn_model.setEnabled(False)
        self.btn_export.setEnabled(False)
        self.table.setRowCount(0)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.video_label.setText("⚙ 正在调用 GPU 极速推理...\n\n请稍候，正在重构物理时序")
        
        self.process_thread = FastProcessThread(file_path, self.model_path)
        self.process_thread.progress.connect(lambda p: self.progress_bar.setValue(p))
        self.process_thread.finished.connect(self.on_analysis_finished)
        self.process_thread.error.connect(self.on_analysis_error)
        self.process_thread.start()

    def on_analysis_finished(self, records, fps, total_frames, video_path, total_count, count_history):
        self.cycle_records = records
        self.fps = fps
        self.total_frames = total_frames
        self.total_count = total_count
        self.count_history = count_history
        
        self.progress_bar.setVisible(False)
        self.btn_load.setEnabled(True)
        self.btn_model.setEnabled(True)
        self.btn_export.setEnabled(True)
        
        self.cap = cv2.VideoCapture(video_path)
        self.slider.setRange(0, self.total_frames - 1)
        self.slider.setValue(0)
        self.btn_play.setEnabled(True)
        
        self.update_leaderboard()
        self.update_frame() 
        QMessageBox.information(self, "分析完成", f"极速分析完毕！\n总完成件数: {self.total_count} 件\n有效生产周期: {len(records)} 个。")

    def on_analysis_error(self, err_msg):
        self.progress_bar.setVisible(False)
        self.btn_load.setEnabled(True)
        self.btn_model.setEnabled(True)
        self.video_label.setText("❌ 分析失败！")
        QMessageBox.critical(self, "错误", f"分析失败:\n{err_msg}")

    def export_report(self):
        if not self.cycle_records:
            QMessageBox.warning(self, "提示", "暂无数据可导出！")
            return

        default_name = f"IE效能报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        fp, _ = QFileDialog.getSaveFileName(self, "导出 Excel 报告", default_name, "Excel 文件 (*.xlsx)")
        if not fp: return

        temp_dir = Path(fp).parent
        line_chart_path = str(temp_dir / "temp_line_chart.png")
        bar_chart_path = str(temp_dir / "temp_bar_chart.png")

        try:
            df = pd.DataFrame(self.cycle_records)
            df['耗时(秒)'] = df['time'].round(2)
            df['时间段'] = df.apply(lambda x: f"{self.format_time(x['start_frame']/self.fps)} - {self.format_time(x['end_frame']/self.fps)}", axis=1)
            
            times = df['time'].values
            normal_time = calculate_standard_time(times)
            allowance_rate = 0.05
            standard_time = normal_time * (1 + allowance_rate)
            compliance_rate = calculate_compliance_rate(times, normal_time)

            # ★ 使用 Matplotlib 生成专业图表 (修复图例与字体) ★
            # 1. 耗时趋势折线图
            plt.figure(figsize=(10, 5))
            line, = plt.plot(range(1, len(times)+1), times, marker='o', color='#4472C4')
            line.set_label('周期耗时')
            axh = plt.axhline(y=normal_time, color='g', linestyle='--', linewidth=2)
            axh.set_label(f'正常工时: {normal_time:.2f}s')
            plt.title('各周期耗时趋势折线图', fontsize=14)
            plt.xlabel('周期序号', fontsize=12)
            plt.ylabel('耗时(秒)', fontsize=12)
            plt.legend(loc='best') # 自动寻找最佳位置显示图例
            plt.grid(True, linestyle='--', alpha=0.6)
            plt.tight_layout()
            plt.savefig(line_chart_path, dpi=150)
            plt.close()

            # 2. 耗时频次分布直方图 (带垂直绿线)
            plt.figure(figsize=(10, 5))
            plt.hist(times, bins=min(10, len(times)), color='#4472C4', edgecolor='white', alpha=0.7)
            # 为直方图添加图例代理
            hist_patch = mpatches.Patch(color='#4472C4', label='频次分布')
            
            axv = plt.axvline(x=normal_time, color='green', linewidth=3, linestyle='--')
            axv.set_label(f'正常工时: {normal_time:.2f}s')
            
            plt.title('耗时频次分布直方图', fontsize=14)
            plt.xlabel('耗时区间(秒)', fontsize=12)
            plt.ylabel('频次', fontsize=12)
            # 明确指定图例的 handles 和 labels
            plt.legend(handles=[hist_patch, axv], loc='best')
            plt.grid(True, linestyle='--', alpha=0.6)
            plt.tight_layout()
            plt.savefig(bar_chart_path, dpi=150)
            plt.close()

            with pd.ExcelWriter(fp, engine='openpyxl') as writer:
                df[['cycle_id', '耗时(秒)', '时间段']].to_excel(writer, sheet_name='明细数据', index=False)
                
                if len(times) > 1:
                    counts, edges = np.histogram(times, bins=min(10, len(times)))
                    hist_data = [{'区间': f"{edges[i]:.2f}-{edges[i+1]:.2f}", '频次': counts[i]} for i in range(len(counts))]
                    hist_df = pd.DataFrame(hist_data)
                else:
                    hist_df = pd.DataFrame({'区间': [f"{times[0]:.2f}"], '频次': [1]})
                hist_df.to_excel(writer, sheet_name='频次分布', index=False)
                
                summary_data = {
                    '指标': ['总完成件数(件)', '有效生产周期数(个)', '正常平均工时(秒)', '宽放率(%)', '标准工时(秒)', '时序符合度(%)', '最大异常工时(秒)', '最小极限工时(秒)'],
                    '数值': [self.total_count, len(times), round(normal_time, 2), f"{allowance_rate * 100:.0f}%", round(standard_time, 2), f"{compliance_rate:.1f}%", round(max(times), 2) if len(times)>0 else 0, round(min(times), 2) if len(times)>0 else 0],
                    '说明': ['系统检测到的原始放置次数', '剔除碎片化误触碰后的真实周期(件数-1)', 'MAD算法剔除异常后的均值', '车间默认生理/疲劳宽放', '正常工时 × (1 + 宽放率)', '耗时在正常工时±20%区间的周期占比', '最长一次周期的耗时', '最短一次周期的耗时']
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='IE效能汇总', index=False)
                
                workbook = writer.book
                ws_detail = writer.sheets['明细数据']
                ws_hist = writer.sheets['频次分布']
                ws_summary = writer.sheets['IE效能汇总']
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                for ws in [ws_detail, ws_hist, ws_summary]:
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    for col_idx, column in enumerate(ws.columns, 1):
                        max_width = 0
                        col_letter = get_column_letter(col_idx)
                        for cell in column:
                            if cell.value is not None:
                                cell_width = get_excel_col_width(cell.value)
                                if cell_width > max_width:
                                    max_width = cell_width
                        ws.column_dimensions[col_letter].width = min(max_width + 2, 60)

                img1 = XLImage(line_chart_path)
                img1.width, img1.height = 720, 360
                ws_detail.add_image(img1, "E2")
                
                img2 = XLImage(bar_chart_path)
                img2.width, img2.height = 720, 360
                ws_hist.add_image(img2, "E2")

            QMessageBox.information(self, "成功", f"报告已成功导出至:\n{fp}\n\n总件数: {self.total_count}\n有效周期: {len(times)}\n标准工时: {standard_time:.2f}秒")
        except Exception as e:
            QMessageBox.critical(self, "失败", f"导出失败:\n{str(e)}")
        finally:
            if Path(line_chart_path).exists(): Path(line_chart_path).unlink()
            if Path(bar_chart_path).exists(): Path(bar_chart_path).unlink()

    def toggle_play(self):
        if not self.cap: return
        if self.timer.isActive():
            self.timer.stop(); self.btn_play.setText("▶ 播放")
        else:
            self.timer.start(int(1000 / self.fps)); self.btn_play.setText("⏸ 暂停")

    def seek(self, frame_idx):
        if not self.cap: return
        self.cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
        self.update_frame()

    def format_time(self, sec):
        m, s = divmod(int(sec), 60)
        return f"{m:02d}:{s:02d}"

    def get_current_count(self, cur_frame):
        count = 0
        for f_idx, c in self.count_history:
            if cur_frame >= f_idx:
                count = c
            else:
                break
        return count

    def update_frame(self):
        if not self.cap or not self.yolo: return
        ret, frame = self.cap.read()
        if not ret:
            self.timer.stop(); self.btn_play.setText("▶ 播放 (结束)")
            return
        
        cur_frame = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES))
        
        results = self.yolo.predict(frame, conf=0.25, iou=0.45, verbose=False)
        boxes = results[0].boxes.data.cpu().numpy()
        
        for b in boxes:
            x1, y1, x2, y2 = map(int, b[:4])
            cls = int(b[5])
            conf = b[4]
            if cls == 0: 
                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.putText(frame, f"HAND {conf:.2f}", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
            elif cls == 1: 
                cv2.rectangle(frame, (x1, y1), (x2, y2), (255, 191, 0), 2)
                cv2.putText(frame, f"BASKET {conf:.2f}", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 191, 0), 2)

        current_count = self.get_current_count(cur_frame)
        self._draw_hud(frame, current_count, cur_frame)
        
        self.slider.blockSignals(True); self.slider.setValue(cur_frame); self.slider.blockSignals(False)
        cur_time = cur_frame / self.fps; total_time = self.total_frames / self.fps
        self.lbl_time.setText(f"{self.format_time(cur_time)} / {self.format_time(total_time)}")
        
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb.shape
        qt_img = QImage(rgb.data, w, h, ch * w, QImage.Format_RGB888).copy()
        self.video_label.setPixmap(QPixmap.fromImage(qt_img).scaled(self.video_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))

    def _draw_hud(self, frame, count, cur_frame):
        h, w = frame.shape[:2]
        color = (0, 240, 255) 
        thickness = 2
        length = 30
        
        cv2.line(frame, (10, 10), (10+length, 10), color, thickness)
        cv2.line(frame, (10, 10), (10, 10+length), color, thickness)
        cv2.line(frame, (w-10, 10), (w-10-length, 10), color, thickness)
        cv2.line(frame, (w-10, 10), (w-10, 10+length), color, thickness)
        cv2.line(frame, (10, h-10), (10+length, h-10), color, thickness)
        cv2.line(frame, (10, h-10), (10, h-10-length), color, thickness)
        cv2.line(frame, (w-10, h-10), (w-10-length, h-10), color, thickness)
        cv2.line(frame, (w-10, h-10), (w-10, h-10-length), color, thickness)
        
        scan_y = int((cur_frame % 60) / 60 * h)
        cv2.line(frame, (10, scan_y), (w-10, scan_y), color, 1)
        
        cv2.putText(frame, f"COUNT: {count}", (40, 50), cv2.FONT_HERSHEY_SIMPLEX, 1.5, (0, 240, 255), 3)
        cv2.putText(frame, "VISION SYSTEM V1.0", (w - 300, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 0), 2)
        cv2.putText(frame, f"FRAME: {cur_frame}/{self.total_frames}", (40, h - 40), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (200, 200, 200), 1)

    def update_leaderboard(self):
        if not self.cycle_records: return
        sorted_records = sorted(self.cycle_records, key=lambda x: x['time'])
        self.table.setRowCount(len(sorted_records))
        
        normal_time = calculate_standard_time([r['time'] for r in self.cycle_records])
        compliance = calculate_compliance_rate([r['time'] for r in self.cycle_records], normal_time)
        self.board_title.setText(f"📊 节拍效能看板 | 总件数: {self.total_count} | 周期数: {len(self.cycle_records)} | 均时: {normal_time:.2f}s | 符合度: {compliance:.1f}%")
        
        for row, record in enumerate(sorted_records):
            rank_item = QTableWidgetItem(str(row + 1))
            cycle_item = QTableWidgetItem(f"周期 {record['cycle_id']}")
            time_item = QTableWidgetItem(f"{record['time']:.2f}")
            start_time_str = self.format_time(record['start_frame'] / self.fps)
            end_time_str = self.format_time(record['end_frame'] / self.fps)
            time_range_item = QTableWidgetItem(f"{start_time_str} - {end_time_str}")
            
            if row == 0: color = QColor(0, 255, 157)   
            elif row == len(sorted_records) - 1: color = QColor(255, 107, 107) 
            else: color = QColor(200, 200, 200)
                
            items = [rank_item, cycle_item, time_item, time_range_item]
            for col, item in enumerate(items):
                item.setTextAlignment(Qt.AlignCenter)
                item.setForeground(color)
                self.table.setItem(row, col, item)

    def on_table_clicked(self, row, col):
        if not self.cap: return
        sorted_records = sorted(self.cycle_records, key=lambda x: x['time'])
        if row < len(sorted_records):
            selected_record = sorted_records[row]
            start_frame = selected_record['start_frame']
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)
            self.slider.setValue(start_frame)
            if not self.timer.isActive(): self.toggle_play()

if __name__ == '__main__':
    crop_dir = Path("/home/ai/data_disk1/yolo/PJ1/crops")
    crop_dir.mkdir(parents=True, exist_ok=True)
    app = QApplication(sys.argv)
    window = DemoApp()
    window.show()
    sys.exit(app.exec())
