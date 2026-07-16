import sys
import cv2
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Tuple, Optional
from enum import Enum, auto
import threading
import queue
from collections import deque
import logging

from PySide6.QtCore import Qt, QTimer, QThread, Signal
from PySide6.QtGui import QImage, QPixmap, QColor
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QSlider, QTableWidget, QTableWidgetItem,
    QHeaderView, QFileDialog, QMessageBox,
    QSplitter, QSizePolicy, QProgressBar, QFrame
)
import onnxruntime as ort

# 屏蔽 ONNX Runtime 底层因环境缺失产生的红字警告
logging.basicConfig(level=logging.ERROR)
ort.set_default_logger_severity(3)

# ═══════════════════════════════════════════════════════════════
# 极致推理引擎 (专为 Windows 异构加速设计)
# ═══════════════════════════════════════════════════════════════

class YOLO_ONNX_Detector:
    def __init__(self, model_path: str):
        available = ort.get_available_providers()
        self.session = None
        self.active_device = "CPU"

        # 优先级 1: CUDA (Windows 下 N 卡加速的绝对主力)
        if 'CUDAExecutionProvider' in available:
            try:
                self.session = ort.InferenceSession(model_path, providers=['CUDAExecutionProvider'])
                self.active_device = "GPU (CUDA)"
            except Exception:
                pass # 如果 CUDA 加载失败，静默跳过，尝试下一个

        # 优先级 2: 纯 CPU 兜底 (保证在任何 Windows 电脑上都能跑)
        if self.session is None:
            self.session = ort.InferenceSession(model_path, providers=['CPUExecutionProvider'])
            self.active_device = "CPU"
            
        self.input_name = self.session.get_inputs()[0].name
        self.imgsz = 640

    def _letterbox(self, img: np.ndarray):
        h, w = img.shape[:2]
        r = min(self.imgsz / h, self.imgsz / w)
        new_h, new_w = int(h * r), int(w * r)
        resized = cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_LINEAR)
        canvas = np.full((self.imgsz, self.imgsz, 3), 114, dtype=np.uint8)
        top = (self.imgsz - new_h) // 2
        left = (self.imgsz - new_w) // 2
        canvas[top:top + new_h, left:left + new_w] = resized
        return canvas, r, (left, top)

    def _preprocess(self, frame: np.ndarray):
        canvas, r, (left, top) = self._letterbox(frame)
        canvas = cv2.cvtColor(canvas, cv2.COLOR_BGR2RGB)
        canvas = canvas.astype(np.float32) / 255.0
        canvas = canvas.transpose(2, 0, 1)[None]  # [1, 3, 640, 640]
        return canvas, r, left, top, frame.shape[:2]

    def _postprocess(self, output: np.ndarray, r: float, left: int, top: int, orig_shape: tuple):
        output = output[0].T  # [8400, 5] or [8400, 6]
        scores = output[:, 4:].max(axis=1)
        mask = scores > 0.5
        
        if not np.any(mask):
            return np.zeros((0, 6), dtype=np.float32)
            
        boxes_xywh = output[mask, :4]
        scores = scores[mask]
        
        boxes_xyxy = np.zeros((len(boxes_xywh), 4), dtype=np.float32)
        boxes_xyxy[:, 0] = boxes_xywh[:, 0] - boxes_xywh[:, 2] / 2
        boxes_xyxy[:, 1] = boxes_xywh[:, 1] - boxes_xywh[:, 3] / 2
        boxes_xyxy[:, 2] = boxes_xywh[:, 0] + boxes_xywh[:, 2] / 2
        boxes_xyxy[:, 3] = boxes_xywh[:, 1] + boxes_xywh[:, 3] / 2
        
        indices = cv2.dnn.NMSBoxes(boxes_xyxy.tolist(), scores.tolist(), 0.5, 0.45)
        if len(indices) == 0:
            return np.zeros((0, 6), dtype=np.float32)
            
        indices = np.array(indices).flatten()
        boxes_xyxy = boxes_xyxy[indices]
        scores = scores[indices]
        
        orig_h, orig_w = orig_shape
        boxes_xyxy[:, [0, 2]] = (boxes_xyxy[:, [0, 2]] - left) / r
        boxes_xyxy[:, [1, 3]] = (boxes_xyxy[:, [1, 3]] - top) / r
        boxes_xyxy[:, [0, 2]] = np.clip(boxes_xyxy[:, [0, 2]], 0, orig_w)
        boxes_xyxy[:, [1, 3]] = np.clip(boxes_xyxy[:, [1, 3]], 0, orig_h)
        
        cls_arr = np.zeros((len(boxes_xyxy), 1), dtype=np.float32)
        return np.concatenate([boxes_xyxy, scores[:, None], cls_arr], axis=1)

    def predict(self, frames: list):
        all_boxes = []
        for frame in frames:
            blob, r, left, top, orig_shape = self._preprocess(frame)
            outputs = self.session.run(None, {self.input_name: blob})[0]
            boxes = self._postprocess(outputs, r, left, top, orig_shape)
            all_boxes.append(boxes)
        return all_boxes

# ═══════════════════════════════════════════════════════════════
# 数据层：不可变事件溯源
# ═══════════════════════════════════════════════════════════════

class ActionStatus(Enum):
    NORMAL = auto()
    ABNORMAL = auto()

@dataclass(frozen=True)
class ActionEvent:
    index: int
    start_time: float
    end_time: float
    duration: float
    status: ActionStatus
    start_frame: int
    end_frame: int

    def to_dict(self) -> dict:
        return {
            'action_index': self.index, 'start_time': self.start_time,
            'end_time': self.end_time, 'duration': self.duration,
            'status': '异常' if self.status == ActionStatus.ABNORMAL else '正常',
            'start_frame': self.start_frame, 'end_frame': self.end_frame
        }

@dataclass
class AnalysisResult:
    events: List[ActionEvent] = field(default_factory=list)
    product_count: int = 0
    total_count: int = 0
    abnormal_count: int = 0
    normal_avg_duration: float = 0.0

    @property
    def normal_events(self) -> List[ActionEvent]:
        return [e for e in self.events if e.status == ActionStatus.NORMAL]

    @property
    def abnormal_events(self) -> List[ActionEvent]:
        return [e for e in self.events if e.status == ActionStatus.ABNORMAL]

    def recompute_stats(self):
        self.total_count = len(self.events)
        self.abnormal_count = len(self.abnormal_events)
        normals = self.normal_events
        self.normal_avg_duration = np.mean([e.duration for e in normals]) if normals else 0.0

# ═══════════════════════════════════════════════════════════════
# 核心引擎：反脆弱自适应状态机
# ═══════════════════════════════════════════════════════════════

class StateMachine:
    def __init__(self, confirm_frames: int = 3, cooldown_sec: float = 1.5,
                 miss_tolerance: int = 5, max_cycle_sec: float = 300.0):
        self.confirm_frames = confirm_frames
        self.cooldown_sec = cooldown_sec
        self.miss_tolerance = miss_tolerance
        self.max_cycle_sec = max_cycle_sec

        self.window_size = 30
        self.k_sigma = 3.0

        self.reset()
        self.current_dyn_min = 0.0
        self.current_dyn_max = self.max_cycle_sec

    def reset(self):
        self.state = 'IDLE'
        self.frame_counter = 0
        self.miss_counter = 0
        self.count = 0
        self.last_action_start_time = None
        self.last_action_start_frame = 0
        self.cooldown_start_time = 0.0
        self.events = []
        self._max_processed_frame = -1

        self.duration_window = deque(maxlen=self.window_size)
        self.current_dyn_min = 0.0
        self.current_dyn_max = self.max_cycle_sec
        self._is_bootstrapped = False

    def _calculate_robust_bounds(self) -> Tuple[float, float]:
        n = len(self.duration_window)
        if n == 0:
            return 0.0, self.max_cycle_sec

        data = np.array(self.duration_window)
        median = np.median(data)
        mad = np.median(np.abs(data - median))
        std_est = max(0.5, 1.4826 * mad)

        if n < 5: k_factor = 1.5
        elif n < 15: k_factor = 2.0
        else: k_factor = 3.0

        dyn_min = max(2.0, median - k_factor * std_est)
        dyn_max = min(self.max_cycle_sec, median + k_factor * std_est)

        self.current_dyn_min = dyn_min
        self.current_dyn_max = dyn_max
        return dyn_min, dyn_max

    def process(self, detected: bool, time_sec: float, frame_idx: int):
        if frame_idx <= self._max_processed_frame:
            return self.state, None

        event = None

        def trigger_new_action():
            nonlocal event
            if self.last_action_start_time is not None:
                interval = time_sec - self.last_action_start_time
                if not self._is_bootstrapped:
                    self.duration_window.extend([interval] * min(10, self.window_size))
                    self._is_bootstrapped = True
                dyn_min, dyn_max = self._calculate_robust_bounds()
                if interval > dyn_max or interval < dyn_min:
                    status = ActionStatus.ABNORMAL
                else:
                    status = ActionStatus.NORMAL
                event = ActionEvent(
                    index=self.count, start_time=self.last_action_start_time,
                    end_time=time_sec, duration=interval, status=status,
                    start_frame=self.last_action_start_frame, end_frame=frame_idx
                )
                self.events.append(event)
                if status == ActionStatus.NORMAL:
                    self.duration_window.append(interval)
            self.count += 1
            self.state = 'LOCKED'
            self.last_action_start_time = time_sec
            self.last_action_start_frame = frame_idx
            self.frame_counter = 0
            self.miss_counter = 0

        if self.state == 'IDLE':
            if detected:
                self.frame_counter += 1
                if self.frame_counter >= self.confirm_frames:
                    trigger_new_action()
            else:
                self.frame_counter = 0

        elif self.state == 'LOCKED':
            if detected:
                self.miss_counter = 0
            else:
                self.miss_counter += 1
                if self.miss_counter > self.miss_tolerance:
                    self.state = 'COOLDOWN'
                    self.cooldown_start_time = time_sec
                    self.frame_counter = 0

        elif self.state == 'COOLDOWN':
            if detected:
                self.frame_counter += 1
                if self.frame_counter >= self.confirm_frames:
                    trigger_new_action()
            elif (time_sec - self.cooldown_start_time) >= self.cooldown_sec:
                self.state = 'IDLE'
                self.frame_counter = 0

        self._max_processed_frame = frame_idx
        return self.state, event

    def global_refine(self):
        if not self.events: return

        durations = np.array([e.duration for e in self.events])
        median = np.median(durations)
        mad = np.median(np.abs(durations - median))
        std_est = max(0.5, 1.4826 * mad)

        k_factor = 3.0
        self.current_dyn_min = max(2.0, median - k_factor * std_est)
        self.current_dyn_max = min(self.max_cycle_sec, median + k_factor * std_est)

        refined_events = []
        ghost_action_count = 0

        for e in self.events:
            is_short_abnormal = (e.duration < self.current_dyn_min)
            is_long_abnormal = (e.duration > self.current_dyn_max)

            if is_short_abnormal:
                new_status = ActionStatus.ABNORMAL
                ghost_action_count += 1
            elif is_long_abnormal:
                new_status = ActionStatus.ABNORMAL
            else:
                new_status = ActionStatus.NORMAL

            refined_events.append(ActionEvent(
                index=e.index, start_time=e.start_time, end_time=e.end_time,
                duration=e.duration, status=new_status,
                start_frame=e.start_frame, end_frame=e.end_frame
            ))

        self.events = refined_events
        self.count = max(0, self.count - ghost_action_count)

    def get_result(self) -> AnalysisResult:
        result = AnalysisResult(events=list(self.events), product_count=self.count)
        result.recompute_stats()
        return result

# ═══════════════════════════════════════════════════════════════
# 视频元数据
# ═══════════════════════════════════════════════════════════════

class VideoMetadata:
    def __init__(self, path: str):
        self.path = path
        cap = cv2.VideoCapture(path)
        if not cap.isOpened():
            raise ValueError(f"无法打开视频: {path}")
        self.fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
        if self.fps <= 0: self.fps = 30.0
        self.total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        self.width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        self.height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self.duration_sec = self.total_frames / self.fps
        cap.release()

    def format_time(self, seconds: float) -> str:
        mins = int(seconds // 60)
        secs = int(seconds % 60)
        return f"{mins:02d}:{secs:02d}"

# ═══════════════════════════════════════════════════════════════
# 绝对刚性死守防抖追踪器
# ═══════════════════════════════════════════════════════════════

class SpatialAnchorTracker:
    def __init__(self):
        self.ref_box = None               
        self.history = []                 
        self.calibration_count = 10       
        self.alpha = 0.1                  
        self.tolerance_ratio = 0.5        
        
    def update(self, boxes_data: np.ndarray) -> Tuple[Optional[np.ndarray], Optional[np.ndarray]]:
        if len(boxes_data) == 0:
            return None, self.ref_box
            
        if self.ref_box is None:
            best_box = max(boxes_data, key=lambda b: b[4])
            self.history.append(best_box[:4])
            if len(self.history) >= self.calibration_count:
                self.ref_box = np.median(np.array(self.history), axis=0)
            return best_box, None

        ref_cx = (self.ref_box[0] + self.ref_box[2]) / 2
        ref_cy = (self.ref_box[1] + self.ref_box[3]) / 2
        
        distances = []
        for box in boxes_data:
            cx = (box[0] + box[2]) / 2
            cy = (box[1] + box[3]) / 2
            dist = np.sqrt((ref_cx - cx)**2 + (ref_cy - cy)**2)
            distances.append(dist)
            
        min_dist_idx = np.argmin(distances)
        best_box = boxes_data[min_dist_idx]
        min_dist = distances[min_dist_idx]
        
        ref_w = self.ref_box[2] - self.ref_box[0]
        ref_h = self.ref_box[3] - self.ref_box[1]
        dist_threshold = max(ref_w, ref_h) * self.tolerance_ratio
        
        if min_dist <= dist_threshold:
            best_xyxy = np.array(best_box[:4], dtype=np.float32)
            self.ref_box = self.alpha * best_xyxy + (1 - self.alpha) * self.ref_box
            return best_box, self.ref_box
        else:
            return None, self.ref_box

# ═══════════════════════════════════════════════════════════════
# 预处理线程
# ═══════════════════════════════════════════════════════════════

class PreprocessThread(QThread):
    progress = Signal(int, str, object)
    finished = Signal(object, list)
    error = Signal(str)

    def __init__(self, video_path: str, model_path: str, batch_size: int = 16):
        super().__init__()
        self.video_path = video_path
        self.model_path = model_path
        self.batch_size = batch_size
        self.sm = StateMachine()
        self.tracker = SpatialAnchorTracker()
        self._is_cancelled = False
        self._frame_queue = queue.Queue(maxsize=512)
        self.detection_history = []
        self.model = None

    def cancel(self):
        self._is_cancelled = True
        while not self._frame_queue.empty():
            try: self._frame_queue.get_nowait()
            except queue.Empty: break
        self._frame_queue.put(None)

    def _read_frames_worker(self):
        cap = cv2.VideoCapture(self.video_path)
        if not cap.isOpened():
            self.error.emit(f"子线程无法打开视频: {self.video_path}")
            self._frame_queue.put(None)
            return
        try:
            idx = 0
            while not self._is_cancelled:
                ret, frame = cap.read()
                if not ret: break
                self._frame_queue.put((idx, frame))
                idx += 1
        except Exception as e:
            self.error.emit(f"视频读取异常: {str(e)}")
        finally:
            self._frame_queue.put(None)
            cap.release()

    def run(self):
        try:
            self.model = YOLO_ONNX_Detector(self.model_path)
            
            meta = VideoMetadata(self.video_path)
            self.detection_history = [None] * meta.total_frames

            reader_thread = threading.Thread(target=self._read_frames_worker, daemon=True)
            reader_thread.start()

            batch_frames, batch_indices = [], []

            while not self._is_cancelled:
                item = self._frame_queue.get()
                if item is None: break
                idx, frame = item
                batch_frames.append(frame)
                batch_indices.append(idx)
                if len(batch_frames) >= self.batch_size:
                    self._process_batch(meta, batch_frames, batch_indices)
                    batch_frames.clear()
                    batch_indices.clear()

            if not self._is_cancelled and batch_frames:
                self._process_batch(meta, batch_frames, batch_indices)

            reader_thread.join()
            if not self._is_cancelled:
                self.sm.global_refine()
                self.finished.emit(self.sm.get_result(), self.detection_history)
        except Exception as e:
            self.error.emit(f"推理主线程异常: {str(e)}")

    def _process_batch(self, meta, batch_frames, batch_indices):
        all_boxes = self.model.predict(frames=batch_frames)
        for i, boxes_data in enumerate(all_boxes):
            valid_box, current_ref = self.tracker.update(boxes_data)
            
            detected = valid_box is not None
            time_sec = batch_indices[i] / meta.fps
            self.sm.process(detected, time_sec, batch_indices[i])

            if detected:
                self.detection_history[batch_indices[i]] = {'box': valid_box.reshape(1, -1), 'ref': current_ref}
            else:
                self.detection_history[batch_indices[i]] = {'box': None, 'ref': current_ref}

        last_idx = batch_indices[-1] + 1
        progress = min(100, int((last_idx / meta.total_frames) * 100)) if meta.total_frames else 100
        time_text = f"{meta.format_time(last_idx/meta.fps)} / {meta.format_time(meta.duration_sec)}"
        self.progress.emit(progress, time_text, self.sm.get_result())

# ═══════════════════════════════════════════════════════════════
# 图表生成器
# ═══════════════════════════════════════════════════════════════

class ChartGenerator:
    @staticmethod
    def setup_chinese_font():
        font_paths = [
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/usr/share/fonts/truetype/msyh/msyh.ttc",
            "C:/Windows/Fonts/msyh.ttc", "C:/Windows/Fonts/simhei.ttf",
            "/System/Library/Fonts/PingFang.ttc", "/Library/Fonts/Arial Unicode.ttf"
        ]
        for fp in font_paths:
            if Path(fp).exists():
                fm.fontManager.addfont(fp)
                font_name = fm.FontProperties(fname=fp).get_name()
                plt.rcParams['font.sans-serif'] = [font_name] + plt.rcParams.get('font.sans-serif', [])
                break
        plt.rcParams['axes.unicode_minus'] = False

    @classmethod
    def generate_report_chart(cls, result: AnalysisResult, path: str):
        cls.setup_chinese_font()
        fig, axes = plt.subplots(2, 1, figsize=(10, 6), dpi=150, facecolor='white')
        report_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        fig.suptitle(f'工业视觉效能透视分析报告 (生成于: {report_time})',
                     fontsize=15, color='#333333', weight='bold', y=0.98)

        try:
            ax1, ax2 = axes
            df = pd.DataFrame([e.to_dict() for e in result.events])
            COLOR_NORMAL, COLOR_ABNORMAL, COLOR_AVG_LINE = '#2E86AB', '#E84855', '#F18F01'

            ax1.set_facecolor('#FAFAFA')
            if not df.empty:
                normal = df[df['status'] == '正常']
                abnormal = df[df['status'] == '异常']
                if not normal.empty:
                    ax1.plot(normal['action_index'], normal['duration'],
                             color=COLOR_NORMAL, marker='o', markersize=3,
                             label='正常节拍', alpha=0.8)
                if not abnormal.empty:
                    ax1.scatter(abnormal['action_index'], abnormal['duration'],
                                color=COLOR_ABNORMAL, s=50, marker='v',
                                label='异常超时/误判', zorder=5,
                                edgecolors='black', linewidths=0.5)
                max_idx = df['action_index'].max()
                max_dur = df['duration'].max()
                ax1.set_xlim(left=-1, right=max_idx + max(2, max_idx * 0.05))
                ax1.set_ylim(bottom=0, top=max_dur * 1.2)

            ax1.set_title('动作耗时趋势监控', color='black', fontsize=11, pad=8)
            ax1.set_xlabel('动作序号', color='#555555')
            ax1.set_ylabel('耗时(秒)', color='#555555')
            ax1.legend(loc='upper right', facecolor='white',
                       edgecolor='#CCCCCC', fontsize=9, framealpha=0.9)
            ax1.grid(True, linestyle='--', alpha=0.4, color='#CCCCCC')

            ax2.set_facecolor('#FAFAFA')
            if not df.empty:
                max_t = int(df['duration'].max()) + 2
                bin_width = max(1, int(max_t / 15))
                bins = list(range(0, max_t + bin_width, bin_width))
                counts, _, _ = ax2.hist(df['duration'], bins=bins, color=COLOR_NORMAL,
                         edgecolor='white', alpha=0.7)
                normal_durations = df[df['status'] == '正常']['duration']
                if not normal_durations.empty:
                    avg = normal_durations.mean()
                    ax2.axvline(avg, color=COLOR_AVG_LINE, linestyle='--',
                                linewidth=2, label=f'正常均值: {avg:.2f}秒')
                max_count = max(counts) if len(counts) > 0 else 1
                ax2.set_ylim(0, max_count * 1.2)

            ax2.set_title('节拍耗时频率分布', color='black', fontsize=11, pad=8)
            ax2.set_xlabel('耗时区间(秒)', color='#555555')
            ax2.set_ylabel('频次', color='#555555')
            ax2.legend(loc='upper right', facecolor='white',
                       edgecolor='#CCCCCC', fontsize=9, framealpha=0.9)
            ax2.grid(True, linestyle='--', alpha=0.4, color='#CCCCCC')

            plt.tight_layout(rect=[0, 0, 1, 0.95])
            plt.savefig(path, facecolor='white', dpi=150, bbox_inches='tight')
        finally:
            plt.close(fig)

# ═══════════════════════════════════════════════════════════════
# 主界面
# ═══════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):
    F_UI = '"Noto Sans CJK SC", "WenQuanYi Micro Hei", "Microsoft YaHei", sans-serif'
    F_MO = '"Noto Sans Mono CJK SC", "Consolas", monospace'

    def __init__(self):
        super().__init__()
        self.setWindowTitle("工业视觉效能透视终端 V13.0 — 极致异构级联架构")
        sg = QApplication.primaryScreen().availableGeometry()
        self.resize(min(2560, max(1200, int(sg.width() * 0.9))),
                    min(1440, max(800, int(sg.height() * 0.9))))
        self.setMinimumSize(1200, 700)

        self._apply_holographic_styles()
        self.model = None
        self.model_path = None
        self.video_meta = None
        self.cap = None
        self.analysis_result = None
        self.preprocess_thread = None
        self.current_result = None
        self.detection_history = []
        
        self.timer = QTimer()
        self.timer.timeout.connect(self._update_frame)
        self.is_playing = False
        self.is_preprocessing = False
        self.replay_sm = StateMachine()
        self._init_ui()
        self._load_model()

    def _apply_holographic_styles(self):
        self.setStyleSheet(f"""
            QMainWindow {{
                background-color: #050810;
                background-image: qradialgradient(cx:0.5, cy:0.5, radius: 1.5,
                    fx:0.5, fy:0.5, stop:0 #0a1525, stop:1 #02040a);
            }}
            QWidget#GlassPanel {{
                background-color: rgba(10, 20, 35, 0.65);
                border: 1px solid rgba(0, 240, 255, 0.15);
                border-radius: 8px;
            }}
            QWidget#EnginePanel {{
                background-color: rgba(13, 20, 35, 0.8);
                border: 1px solid rgba(0, 240, 255, 0.25);
                border-radius: 4px;
            }}
            QSplitter::handle {{ background-color: transparent; }}
            QSplitter::handle:horizontal {{
                width: 2px; margin: 4px 1px;
                border-left: 1px solid rgba(0, 240, 255, 0.1);
                border-right: 1px solid rgba(0, 240, 255, 0.1);
            }}
            QSplitter::handle:vertical {{
                height: 2px; margin: 1px 4px;
                border-top: 1px solid rgba(0, 240, 255, 0.1);
                border-bottom: 1px solid rgba(0, 240, 255, 0.1);
            }}
            QPushButton {{
                background-color: rgba(0, 240, 255, 0.05); color: #00f0ff;
                border: 1px solid rgba(0, 240, 255, 0.4); padding: 8px 18px;
                font-size: 13px; font-weight: bold; border-radius: 4px;
                font-family: {self.F_UI};
            }}
            QPushButton:hover {{ background-color: rgba(0, 240, 255, 0.15);
                border: 1px solid #00f0ff; color: #ffffff; }}
            QPushButton:pressed {{ background-color: rgba(0, 240, 255, 0.3); color: #050810; }}
            QPushButton:disabled {{ background-color: #0d1117; color: #2a3a4a; border-color: #1e2d3d; }}
            QPushButton#ModelBtn {{
                background-color: rgba(255, 165, 0, 0.08); color: #ffa500;
                border: 1px solid rgba(255, 165, 0, 0.5);
            }}
            QPushButton#ModelBtn:hover {{ background-color: rgba(255, 165, 0, 0.2);
                border: 1px solid #ffa500; color: #ffffff; }}
            QPushButton#ModelBtn:pressed {{ background-color: rgba(255, 165, 0, 0.35); color: #050810; }}
            QSlider::groove:horizontal {{
                height: 4px; background: rgba(255, 255, 255, 0.08); border-radius: 2px;
            }}
            QSlider::handle:horizontal {{
                width: 16px; height: 16px; margin: -7px 0;
                background: #050810; border: 2px solid #00f0ff; border-radius: 8px;
            }}
            QSlider::sub-page:horizontal {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #005f6b, stop:1 #00f0ff); border-radius: 2px;
            }}
            QTableWidget {{
                background-color: transparent; color: #c9d1d9;
                gridline-color: rgba(255, 255, 255, 0.05);
                border: none; font-family: {self.F_MO};
                alternate-background-color: rgba(0, 240, 255, 0.03);
                selection-background-color: rgba(0, 240, 255, 0.1);
                selection-color: #00f0ff;
            }}
            QTableWidget::item {{
                background-color: transparent;
                border-bottom: 1px solid rgba(0, 240, 255, 0.05); padding: 4px 4px;
            }}
            QTableWidget::item:hover {{ background-color: rgba(0, 240, 255, 0.08); }}
            QHeaderView::section {{
                background-color: transparent; color: #58a6ff;
                border: none; border-bottom: 1px solid rgba(0, 240, 255, 0.3);
                font-weight: bold; font-family: {self.F_MO}; padding: 8px 4px;
            }}
            QProgressBar {{
                border: 1px solid rgba(0, 240, 255, 0.2); border-radius: 3px;
                text-align: center; background: rgba(13, 17, 23, 0.8);
                color: #00f0ff; height: 20px;
            }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #005f6b, stop:1 #00f0ff); border-radius: 2px;
            }}
            QLabel {{ color: #e0e6ed; background: transparent; }}
        """)

    def _init_ui(self):
        cw = QWidget()
        self.setCentralWidget(cw)
        root = QVBoxLayout(cw)
        sp = 10 if self.width() >= 1920 else 6
        mg = 16 if self.width() >= 1920 else 10
        root.setSpacing(sp)
        root.setContentsMargins(mg, mg, mg, mg)

        ctrl = QHBoxLayout()
        ctrl.setSpacing(10)

        self.btn_model = QPushButton("🧠 导入模型")
        self.btn_model.setObjectName("ModelBtn")
        self.btn_model.setCursor(Qt.PointingHandCursor)
        self.btn_model.setToolTip("导入 ONNX 检测模型")
        self.btn_model.clicked.connect(self._import_model)

        self.btn_load = QPushButton("📁 导入视频")
        self.btn_load.setCursor(Qt.PointingHandCursor)
        self.btn_load.clicked.connect(self._load_video)

        self.btn_play = QPushButton("▶ 播放")
        self.btn_play.setCursor(Qt.PointingHandCursor)
        self.btn_play.clicked.connect(self._toggle_play)
        self.btn_play.setEnabled(False)

        self.btn_export = QPushButton("📊 导出报告")
        self.btn_export.setCursor(Qt.PointingHandCursor)
        self.btn_export.clicked.connect(self._export_report)
        self.btn_export.setEnabled(False)

        self.slider = QSlider(Qt.Horizontal)
        self.slider.setEnabled(False)
        self.slider.sliderReleased.connect(self._on_slider_released)
        self.slider.valueChanged.connect(self._on_slider_moved)

        self.lbl_time = QLabel("00:00 / 00:00")
        time_style = (f"font-size:14px; color:#00f0ff; font-family:{self.F_MO};"
                      f"background:rgba(13, 17, 23, 0.8); padding:4px 12px;"
                      f"border:1px solid rgba(0, 240, 255, 0.2); border-radius:3px;")
        self.lbl_time.setStyleSheet(time_style)
        self.lbl_time.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        ctrl.addWidget(self.btn_model)
        ctrl.addWidget(self.btn_load)
        ctrl.addWidget(self.btn_play)
        ctrl.addWidget(self.btn_export)
        ctrl.addWidget(self.slider, stretch=1)
        ctrl.addWidget(self.lbl_time)
        root.addLayout(ctrl)

        self.lbl_status = QLabel(" 系统就绪 — 请导入 ONNX 模型与视频开始分析")
        status_style = (f"font-size:12px; color:#58a6ff; background:rgba(13, 20, 35, 0.6);"
                        f"padding:6px 12px; border-left:3px solid #58a6ff;"
                        f"font-family:{self.F_MO}; border-radius:2px;")
        self.lbl_status.setStyleSheet(status_style)
        root.addWidget(self.lbl_status)

        self.main_splitter = QSplitter(Qt.Vertical)
        self.video_splitter = QSplitter(Qt.Horizontal)
        self.lbl_original = QLabel("原始画面")
        self.lbl_annotated = QLabel("AI 识别")
        hud_style = (f"background:#000; font-size:16px; color:#2a3a4a;"
                     f"font-family:{self.F_MO};"
                     f"border-top: 2px solid rgba(0, 240, 255, 0.4);"
                     f"border-bottom: 2px solid rgba(0, 240, 255, 0.4);"
                     f"border-left: 1px solid rgba(0, 240, 255, 0.1);"
                     f"border-right: 1px solid rgba(0, 240, 255, 0.1);")
        for lbl in [self.lbl_original, self.lbl_annotated]:
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            lbl.setMinimumSize(320, 180)
            lbl.setStyleSheet(hud_style)

        self.video_splitter.addWidget(self.lbl_original)
        self.video_splitter.addWidget(self.lbl_annotated)
        self.video_splitter.setStretchFactor(0, 1)
        self.video_splitter.setStretchFactor(1, 1)
        self.main_splitter.addWidget(self.video_splitter)

        data_widget = QWidget()
        data_layout = QHBoxLayout(data_widget)
        data_layout.setSpacing(sp)
        data_layout.setContentsMargins(0, 4, 0, 0)

        left = QVBoxLayout()
        left.setSpacing(10)

        stats_frame = QFrame()
        stats_frame.setObjectName("GlassPanel")
        stats_layout = QVBoxLayout(stats_frame)
        stats_layout.setSpacing(12)
        stats_layout.setContentsMargins(12, 18, 12, 12)

        core_widget = QWidget()
        core_layout = QVBoxLayout(core_widget)
        core_layout.setSpacing(2)
        core_layout.setAlignment(Qt.AlignCenter)
        core_widget.setStyleSheet("""
            QWidget {
                background: qradialgradient(cx:0.5, cy:0.5, radius: 0.7,
                    fx:0.5, fy:0.5, stop:0 rgba(255, 215, 0, 0.15), stop:1 rgba(10, 20, 35, 0));
                border-radius: 6px;
            }
        """)

        self.lbl_product_count = QLabel("0")
        self.lbl_product_count.setStyleSheet(
            f"font-size:52px; font-weight:900; color:#FFD700;"
            f"font-family:{self.F_MO}; border:none;")
        self.lbl_product_count.setAlignment(Qt.AlignCenter)

        lbl_product_tag = QLabel("产 品 总 数")
        lbl_product_tag.setStyleSheet(
            f"font-size:14px; color:#8b949e; font-family:{self.F_UI};"
            f"letter-spacing: 6px; border:none;")
        lbl_product_tag.setAlignment(Qt.AlignCenter)

        core_layout.addWidget(self.lbl_product_count)
        core_layout.addWidget(lbl_product_tag)
        stats_layout.addWidget(core_widget)

        sub_stat_widget = QWidget()
        sub_stat_layout = QHBoxLayout(sub_stat_widget)
        sub_stat_layout.setSpacing(8)
        sub_stat_layout.setContentsMargins(0, 6, 0, 0)

        def create_sub_stat(label_text, color):
            w = QWidget()
            w.setStyleSheet(
                f"background: rgba(13, 20, 35, 0.5);"
                f"border: 1px solid {color}30; border-radius: 4px;")
            lay = QVBoxLayout(w)
            lay.setSpacing(2)
            lay.setContentsMargins(4, 6, 4, 6)
            lay.setAlignment(Qt.AlignCenter)

            val = QLabel("0")
            val.setStyleSheet(
                f"font-size:20px; font-weight:bold; color:{color};"
                f"font-family:{self.F_MO};")
            val.setAlignment(Qt.AlignCenter)

            tag = QLabel(label_text)
            tag.setStyleSheet(
                f"font-size:10px; color:#8b949e; letter-spacing: 2px;")
            tag.setAlignment(Qt.AlignCenter)

            lay.addWidget(val)
            lay.addWidget(tag)
            return w, val

        w_cycle, self.lbl_count = create_sub_stat("动作周期", "#00ff9d")
        w_abn, self.lbl_abnormal = create_sub_stat("异常次数", "#ff6b6b")
        w_avg, self.lbl_avg = create_sub_stat("平均周期", "#00f0ff")

        sub_stat_layout.addWidget(w_cycle)
        sub_stat_layout.addWidget(w_abn)
        sub_stat_layout.addWidget(w_avg)
        stats_layout.addWidget(sub_stat_widget)

        self.lbl_progress = QLabel("预处理进度: 0%")
        self.lbl_progress.setStyleSheet(
            f"font-size:14px; color:#ffa500; font-family:{self.F_MO};")
        self.lbl_progress.setVisible(False)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)

        stats_layout.addWidget(self.lbl_progress)
        stats_layout.addWidget(self.progress_bar)

        engine_panel = QFrame()
        engine_panel.setObjectName("EnginePanel")
        engine_layout = QVBoxLayout(engine_panel)
        engine_layout.setContentsMargins(8, 4, 8, 4)
        self.lbl_engine_status = QLabel("⚙ 异构引擎: 待机中...")
        self.lbl_engine_status.setStyleSheet(
            f"font-size:12px; color:#8b949e; font-family:{self.F_MO}; border:none;")
        self.lbl_engine_status.setAlignment(Qt.AlignCenter)
        engine_layout.addWidget(self.lbl_engine_status)
        stats_layout.addWidget(engine_panel)

        left.addWidget(stats_frame)
        left.addStretch()
        data_layout.addLayout(left, stretch=0)

        rank_container = QFrame()
        rank_container.setObjectName("GlassPanel")
        rank_v_layout = QVBoxLayout(rank_container)
        rank_v_layout.setSpacing(4)
        rank_v_layout.setContentsMargins(8, 8, 8, 8)

        title_lbl = QLabel("📊 全局效能排行榜 (按耗时升序)")
        title_lbl.setStyleSheet(
            f"font-size:13px; font-weight:bold; color:#58a6ff;"
            f"border-bottom:1px solid #58a6ff; padding-bottom:4px; background:transparent;")
        rank_v_layout.addWidget(title_lbl)

        self.table_rank = QTableWidget(0, 5)
        self.table_rank.setHorizontalHeaderLabels(
            ["排名", "动作序号", "耗时(秒)", "状态", "发生时间"])
        self.table_rank.horizontalHeader().setDefaultAlignment(
            Qt.AlignCenter | Qt.AlignVCenter)
        self.table_rank.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_rank.verticalHeader().setVisible(False)
        self.table_rank.verticalHeader().setDefaultSectionSize(28)
        self.table_rank.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_rank.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_rank.setShowGrid(False)
        self.table_rank.setAlternatingRowColors(True)
        self.table_rank.cellClicked.connect(self._jump_to_event)
        rank_v_layout.addWidget(self.table_rank)

        data_layout.addWidget(rank_container, stretch=1)

        self.main_splitter.addWidget(data_widget)
        self.main_splitter.setSizes(
            [int(self.height() * 0.55), int(self.height() * 0.45)])
        self.main_splitter.setStretchFactor(0, 1)
        self.main_splitter.setStretchFactor(1, 1)
        root.addWidget(self.main_splitter, stretch=1)

    def _load_model(self):
        paths = ["best.onnx", "model/best.onnx", "weights/best.onnx"]
        for p in paths:
            if Path(p).exists():
                try:
                    self.model_path = p
                    self._update_model_btn(Path(p).stem)
                    self._set_status(f"默认模型已加载: {Path(p).name}", "success")
                    return True
                except Exception:
                    continue
        self._set_status("⚠ 未找到默认模型，请点击 [🧠 导入模型] 选择 .onnx 文件", "warning")
        return False

    def _import_model(self):
        if self.is_preprocessing:
            QMessageBox.information(self, "操作受限", "正在预处理中，请等待完成后再切换模型。")
            return
        if self.is_playing:
            self.timer.stop()
            self.is_playing = False
            self.btn_play.setText("▶ 播放")

        path, _ = QFileDialog.getOpenFileName(
            self, "选择 ONNX 检测模型", "",
            "ONNX 模型文件 (*.onnx)")
        if not path: return
        if not Path(path).exists():
            QMessageBox.critical(self, "文件不存在", f"路径无效:\n{path}")
            return

        try:
            self._set_status(f"正在加载模型: {Path(path).name} ...", "warning")
            QApplication.processEvents()
            
            self.model_path = path
            self._update_model_btn(Path(path).stem)
            self._set_status(f"✓ 模型加载成功: {Path(path).name}", "success")
            self._reset_video_state()
        except Exception as e:
            self._set_status(f"✗ 模型加载失败: {str(e)[:80]}", "error")
            QMessageBox.critical(self, "模型加载失败", f"错误详情:\n{str(e)}")

    def _update_model_btn(self, model_name: str):
        display = model_name[:14] + "…" if len(model_name) > 14 else model_name
        self.btn_model.setText(f"🧠 {display}")
        self.btn_model.setToolTip(f"当前模型: {model_name}\n点击切换其他模型")

    def _reset_video_state(self):
        if self.cap: self.cap.release(); self.cap = None
        self.video_meta = None
        self.analysis_result = None
        self.current_result = None
        self.detection_history = []
        self.replay_sm.reset()
        self._update_stats_ui(AnalysisResult())
        self._clear_tables()
        self.slider.setValue(0)
        self.slider.setEnabled(False)
        self.btn_play.setEnabled(False)
        self.btn_export.setEnabled(False)
        self.lbl_original.setText("原始画面")
        self.lbl_annotated.setText("AI 识别")
        self.lbl_time.setText("00:00 / 00:00")

    def _set_status(self, text, level="info"):
        colors = {"info": ("#58a6ff", "#58a6ff"), "success": ("#00ff9d", "#00ff9d"),
                  "warning": ("#ffa500", "#ffa500"), "error": ("#ff6b6b", "#ff6b6b")}
        c, bc = colors.get(level, colors["info"])
        self.lbl_status.setText(f" {text}")
        self.lbl_status.setStyleSheet(
            f"font-size:12px; color:{c}; background:rgba(13, 20, 35, 0.6);"
            f"padding:6px 12px; border-left:3px solid {bc};"
            f"font-family:{self.F_MO}; border-radius:2px;")

    def _load_video(self):
        if not self.model_path:
            QMessageBox.warning(self, "警告", "模型未加载，请先点击 [🧠 导入模型]")
            return
        if self.is_preprocessing:
            QMessageBox.information(self, "提示", "正在预处理中，请等待完成")
            return

        path, _ = QFileDialog.getOpenFileName(
            self, "选择视频源", "", "视频文件 (*.mp4 *.avi *.mov *.mkv *.wmv)")
        if not path: return

        if self.is_playing:
            self.timer.stop()
            self.is_playing = False
            self.btn_play.setText("▶ 播放")
        if self.cap:
            self.cap.release()
            self.cap = None

        try:
            self.video_meta = VideoMetadata(path)
        except ValueError as e:
            QMessageBox.critical(self, "错误", str(e))
            return

        self.cap = cv2.VideoCapture(path)
        self.slider.setRange(0, self.video_meta.total_frames - 1)
        self.slider.setValue(0)
        self.slider.setEnabled(False)
        self.btn_play.setEnabled(False)
        self.btn_export.setEnabled(False)

        self.analysis_result = None
        self.current_result = None
        self.detection_history = []
        self.replay_sm.reset()
        self._update_stats_ui(AnalysisResult())
        self._clear_tables()

        self.is_preprocessing = True
        self.lbl_original.setText("异构级联加速中...")
        self.lbl_annotated.setText("AI 推理中...")
        self._set_status("正在调用 ONNX Runtime 极速预处理...", "warning")

        self.lbl_progress.setVisible(True)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.lbl_engine_status.setText("⚙ 引擎: 绝对刚性锚定与流式预判中...")
        self.lbl_engine_status.setStyleSheet(
            f"font-size:12px; color:#ffa500; font-family:{self.F_MO}; border:none;")

        self.preprocess_thread = PreprocessThread(path, self.model_path)
        self.preprocess_thread.progress.connect(self._on_preprocess_progress)
        self.preprocess_thread.finished.connect(self._on_preprocess_finished)
        self.preprocess_thread.error.connect(self._on_preprocess_error)
        self.preprocess_thread.start()

    def _on_preprocess_progress(self, progress, time_text, result):
        self.progress_bar.setValue(progress)
        self.lbl_progress.setText(f"预处理进度: {progress}%")
        self.lbl_time.setText(f"[ {progress}% ] {time_text}")
        self.current_result = result
        self._update_stats_ui(result)

        # 获取当前真正使用的硬件设备
        device_name = getattr(self.preprocess_thread.model, 'active_device', 'CPU')

        if self.preprocess_thread.sm._is_bootstrapped:
            min_b = self.preprocess_thread.sm.current_dyn_min
            max_b = self.preprocess_thread.sm.current_dyn_max
            win_size = len(self.preprocess_thread.sm.duration_window)
            anchor_status = "已焊死" if self.preprocess_thread.tracker.ref_box is not None else "锚定中..."
            self.lbl_engine_status.setText(
                f"⚙ 引擎 [{device_name}]: 锚点{anchor_status} | 窗口:{win_size} | 边界: {min_b:.1f}s-{max_b:.1f}s")
            # 如果是 GPU 就显示青色，如果是 CPU 就显示橙色
            color = "#00ff9d" if "GPU" in device_name else "#ffa500"
            self.lbl_engine_status.setStyleSheet(
                f"font-size:12px; color:{color}; font-family:{self.F_MO}; border:none;")
        else:
            self.lbl_engine_status.setText(f"⚙ 引擎 [{device_name}]: 绝对刚性锚定与流式预判中...")
            self.lbl_engine_status.setStyleSheet(
                f"font-size:12px; color:#ffa500; font-family:{self.F_MO}; border:none;")

    def _on_preprocess_finished(self, result, detection_history):
        self.is_preprocessing = False
        self.analysis_result = result
        self.current_result = result
        self.detection_history = detection_history
        self.slider.setEnabled(True)
        self.slider.setValue(0)
        self.btn_play.setEnabled(True)
        self.btn_export.setEnabled(True)
        self.lbl_progress.setVisible(False)
        self.progress_bar.setVisible(False)
        self.lbl_original.setText("审查回放模式")
        self.lbl_annotated.setText("零延迟回放已激活")
        self._set_status(
            f"全局终审完成！产品总数: {result.product_count}，异常 {result.abnormal_count} 次", "success")
        self.lbl_time.setText(f"00:00 / {self.video_meta.format_time(self.video_meta.duration_sec)}")
        self._update_stats_ui(result)
        self._update_tables(result)

        min_b = self.preprocess_thread.sm.current_dyn_min
        max_b = self.preprocess_thread.sm.current_dyn_max
        self.lbl_engine_status.setText(f"⚙ 全局终审已收敛 (精准边界: {min_b:.1f}s - {max_b:.1f}s)")
        self.lbl_engine_status.setStyleSheet(
            f"font-size:12px; color:#58a6ff; font-family:{self.F_MO}; border:none;")

    def _on_preprocess_error(self, error_msg):
        self.is_preprocessing = False
        self.lbl_progress.setVisible(False)
        self.progress_bar.setVisible(False)
        self._set_status(f"预处理失败: {error_msg}", "error")
        QMessageBox.critical(self, "预处理失败", f"错误详情:\n{error_msg}")

    def _toggle_play(self):
        if not self.cap or self.is_preprocessing: return
        self.is_playing = not self.is_playing
        self.btn_play.setText("⏸ 暂停" if self.is_playing else "▶ 播放")
        if self.is_playing: self.timer.start(int(1000 / self.video_meta.fps))
        else: self.timer.stop()

    def _draw_detection_boxes(self, img, history_data):
        ref_box = history_data.get('ref') if history_data else None
        boxes_data = history_data.get('box') if history_data else None

        if ref_box is not None:
            rx1, ry1, rx2, ry2 = map(int, ref_box[:4])
            overlay = img.copy()
            cv2.rectangle(overlay, (rx1, ry1), (rx2, ry2), (255, 191, 0), 2)
            cv2.addWeighted(overlay, 0.6, img, 0.4, 0, img)
            cv2.putText(img, "Rigid Anchor", (rx1, ry1 - 8),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 191, 0), 1, cv2.LINE_AA)

        if boxes_data is not None:
            for box in boxes_data:
                x1, y1, x2, y2 = map(int, box[:4])
                conf = box[4]
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                label = f"{conf:.2f}"
                (w, h), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)
                cv2.rectangle(img, (x1, y1 - h - 5), (x1 + w, y1), (0, 255, 0), -1)
                cv2.putText(img, label, (x1, y1 - 5),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
        return img

    def _update_frame(self):
        if not self.cap or not self.is_playing: return
        ret, frame = self.cap.read()
        if not ret:
            self.timer.stop()
            self.is_playing = False
            self.btn_play.setText("▶ 播放 (已结束)")
            return

        cur_frame = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
        if cur_frame < 0: cur_frame = 0
        cur_time = cur_frame / self.video_meta.fps

        self.slider.blockSignals(True)
        self.slider.setValue(cur_frame)
        self.slider.blockSignals(False)

        self.lbl_time.setText(
            f"{self.video_meta.format_time(cur_time)} / "
            f"{self.video_meta.format_time(self.video_meta.duration_sec)}")

        annotated = frame.copy()
        detected = False
        history_data = self.detection_history[cur_frame] if 0 <= cur_frame < len(self.detection_history) else None
        
        if history_data and history_data.get('box') is not None:
            detected = True
            
        annotated = self._draw_detection_boxes(annotated, history_data)

        state, _ = self.replay_sm.process(detected, cur_time, cur_frame)
        self._draw_holographic_overlay(annotated, cur_frame, state, detected)
        self._show_frame(frame, annotated)

    def _draw_holographic_overlay(self, img, frame_idx, state, detected):
        count = self.replay_sm.count
        overlay = img.copy()
        cv2.rectangle(overlay, (15, 15), (350, 145), (5, 10, 20), -1)
        cv2.addWeighted(overlay, 0.6, img, 0.4, 0, img)

        cv2.putText(img, f"PRODUCT: {count}", (30, 50),
                    cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 240, 255), 2, cv2.LINE_AA)

        status_text = f"STATE: {state}"
        color = ((0, 255, 157) if state == 'IDLE'
                 else (0, 165, 255) if state == 'LOCKED'
                 else (255, 165, 0))
        cv2.putText(img, status_text, (30, 90),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2, cv2.LINE_AA)

        det_text = "DETECTED" if detected else "NO TARGET"
        det_color = (0, 255, 0) if detected else (128, 128, 128)
        cv2.putText(img, det_text, (30, 130),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, det_color, 2, cv2.LINE_AA)

    def _on_slider_moved(self, value):
        if not self.video_meta: return
        t = value / self.video_meta.fps
        self.lbl_time.setText(
            f"{self.video_meta.format_time(t)} / "
            f"{self.video_meta.format_time(self.video_meta.duration_sec)}")

    def _on_slider_released(self):
        if not self.cap or not self.video_meta: return
        pos = max(0, min(self.slider.value(), self.video_meta.total_frames - 1))
        self.cap.set(cv2.CAP_PROP_POS_FRAMES, pos)
        self.replay_sm.reset()
        if not self.is_playing:
            ret, frame = self.cap.read()
            if ret:
                annotated = frame.copy()
                history_data = self.detection_history[pos] if 0 <= pos < len(self.detection_history) else None
                detected = history_data is not None and history_data.get('box') is not None
                annotated = self._draw_detection_boxes(annotated, history_data)
                self._draw_holographic_overlay(annotated, pos, self.replay_sm.state, detected)
                self._show_frame(frame, annotated)

    def _jump_to_event(self, row, col):
        if not self.cap or not self.video_meta: return
        table = self.sender()
        item = table.item(row, 4)
        if not item: return
        try:
            time_str = item.text().split(" - ")[0]
            mins, secs = map(int, time_str.split(":"))
            target_frame = int((mins * 60 + secs) * self.video_meta.fps)
            target_frame = max(0, min(target_frame, self.video_meta.total_frames - 1))
            self.slider.setValue(target_frame)
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, target_frame)
            self.replay_sm.reset()
            ret, frame = self.cap.read()
            if ret:
                annotated = frame.copy()
                history_data = self.detection_history[target_frame] if 0 <= target_frame < len(self.detection_history) else None
                detected = history_data is not None and history_data.get('box') is not None
                annotated = self._draw_detection_boxes(annotated, history_data)
                self._draw_holographic_overlay(annotated, target_frame, self.replay_sm.state, detected)
                self._show_frame(frame, annotated)
        except (ValueError, IndexError): pass

    def _show_frame(self, original, annotated):
        for img, lbl in [(original, self.lbl_original),
                         (annotated, self.lbl_annotated)]:
            rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb.shape
            qt_img = QImage(rgb.data, w, h, ch * w, QImage.Format_RGB888).copy()
            pixmap = QPixmap.fromImage(qt_img)
            lbl.setPixmap(pixmap.scaled(
                lbl.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))

    def _update_stats_ui(self, result):
        self.lbl_product_count.setText(f"{result.product_count}")
        self.lbl_count.setText(f"{result.total_count}")
        self.lbl_abnormal.setText(f"{result.abnormal_count}")
        self.lbl_avg.setText(f"{result.normal_avg_duration:.2f}")

    def _update_tables(self, result):
        sorted_events = sorted(result.events, key=lambda x: x.duration)
        self.table_rank.setRowCount(0)
        for i, evt in enumerate(sorted_events):
            self.table_rank.insertRow(i)
            items = [
                str(i + 1), str(evt.index), f"{evt.duration:.2f}",
                "异常" if evt.status == ActionStatus.ABNORMAL else "正常",
                f"{self.video_meta.format_time(evt.start_time)} - "
                f"{self.video_meta.format_time(evt.end_time)}"
            ]
            for j, txt in enumerate(items):
                it = QTableWidgetItem(txt)
                it.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                if evt.status == ActionStatus.ABNORMAL:
                    it.setForeground(QColor("#ff6b6b"))
                    it.setBackground(QColor(255, 107, 107, 25))
                else:
                    it.setForeground(QColor("#c9d1d9"))
                self.table_rank.setItem(i, j, it)

    def _clear_tables(self):
        self.table_rank.setRowCount(0)

    def _export_report(self):
        if not self.analysis_result or not self.analysis_result.events:
            QMessageBox.information(self, "提示", "无数据可导出！")
            return

        default_name = f"效能分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        fp, _ = QFileDialog.getSaveFileName(
            self, "导出报告", default_name, "Excel 文件 (*.xlsx)")
        if not fp: return

        result = self.analysis_result
        temp_img = str(Path(fp).with_suffix('.png'))

        try:
            ChartGenerator.generate_report_chart(result, temp_img)

            def fmt_sec(sec: float) -> str:
                h, rem = divmod(int(sec), 3600)
                m, s = divmod(rem, 60)
                return f"{h:02d}:{m:02d}:{s:02d}" if h > 0 else f"{m:02d}:{s:02d}"

            detail_rows = []
            abnormal_indices = {
                i + 2 for i, e in enumerate(result.events)
                if e.status == ActionStatus.ABNORMAL
            }

            for e in result.events:
                detail_rows.append({
                    '动作序号': e.index,
                    '开始时间': fmt_sec(e.start_time),
                    '结束时间': fmt_sec(e.end_time),
                    '耗时(秒)': e.duration,
                    '状态': '异常' if e.status == ActionStatus.ABNORMAL else '正常',
                    '起始帧': e.start_frame,
                    '结束帧': e.end_frame
                })
            df_detail = pd.DataFrame(detail_rows)

            summary_data = {
                '指标': ['产品总数', '动作周期数', '异常行为次数',
                         '正常节拍数', '正常平均周期(秒)'],
                '数值': [result.product_count, result.total_count,
                         result.abnormal_count,
                         result.total_count - result.abnormal_count,
                         round(result.normal_avg_duration, 2)]
            }
            df_summary = pd.DataFrame(summary_data)

            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9'))
            center_align = Alignment(horizontal='center', vertical='center')
            header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell_font = Font(name='微软雅黑', size=10)
            err_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            err_font = Font(name='微软雅黑', size=10, color='9C0006', bold=True)

            def get_exact_width(value) -> float:
                if not value: return 6.0
                length = 0.0
                for char in str(value):
                    length += 2.0 if '\u4e00' <= char <= '\u9fff' else 1.1
                return length + 4.0

            with pd.ExcelWriter(fp, engine='openpyxl') as writer:
                df_summary.to_excel(writer, sheet_name='汇总数据', index=False, startrow=1)
                df_detail.to_excel(writer, sheet_name='动作明细', index=False)

                wb = writer.book
                ws_sum = wb['汇总数据']
                ws_sum.merge_cells('A1:B1')
                title_cell = ws_sum['A1']
                title_cell.value = f"效能分析汇总 (生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')})"
                title_cell.font = Font(name='微软雅黑', bold=True, size=13, color='1F4E79')
                title_cell.alignment = Alignment(horizontal='left', vertical='center')
                ws_sum.row_dimensions[1].height = 30

                for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row, max_col=2):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = center_align
                        if cell.row == 2:
                            cell.font = header_font
                            cell.fill = header_fill
                        else:
                            cell.font = cell_font
                    col_letter = row[0].column_letter
                    ws_sum.column_dimensions[col_letter].width = max(
                        get_exact_width(row[0].value), get_exact_width(row[1].value))

                ws_det = wb['动作明细']
                ws_det.freeze_panes = 'A2'
                for cell in ws_det[1]:
                    ws_det.column_dimensions[cell.column_letter].width = get_exact_width(cell.value)

                for row in ws_det.iter_rows(min_row=2, max_row=ws_det.max_row, max_col=ws_det.max_column):
                    is_err_row = row[0].row in abnormal_indices
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = center_align
                        if is_err_row:
                            cell.fill = err_fill
                            cell.font = err_font
                        else:
                            cell.font = cell_font

                ws_chart = wb.create_sheet("可视化图表")
                img = XLImage(temp_img)
                img.width = 960
                img.height = 580
                ws_chart.add_image(img, 'A1')

            QMessageBox.information(self, "导出成功", f"专业级分析报告已保存至:\n{fp}")

        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"错误详情:\n{str(e)}")
        finally:
            if Path(temp_img).exists():
                Path(temp_img).unlink()

    def closeEvent(self, event):
        if self.preprocess_thread and self.preprocess_thread.isRunning():
            self.preprocess_thread.cancel()
            self.preprocess_thread.wait(3000)
        if self.cap:
            self.cap.release()
        event.accept()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
